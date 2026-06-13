"""Document / worksheet / CSV generation, document styles, and resource
browsing tools.

Pure-move of whole functions out of the former single-file module; bodies
are byte-identical.
"""
import os
import csv
import json
import logging
from urllib.parse import quote

from backend.services.assistant_tools import _load_settings, DOCUMENTS_DIR
from backend.utils.compliance import require_teacher_id
from backend.paths import graider_export_dir

from ._paths import PROJECT_ROOT

_logger = logging.getLogger(__name__)

MAX_RESOURCE_TEXT = 120000


def _extract_pdf_text(filepath):
    """Extract text from a PDF file path using PyMuPDF."""
    try:
        import fitz
        doc = fitz.open(filepath)
        pages = []
        for page in doc:
            pages.append(page.get_text())
        page_count = len(pages)
        doc.close()
        return "\n\n".join(pages), page_count
    except ImportError:
        return "[PDF extraction requires PyMuPDF: pip install pymupdf]", 0
    except Exception as e:  # noqa: BLE001  # broad catch: returns fallback
        return f"[Error extracting PDF: {e}]", 0


def _extract_docx_text(filepath):
    """Extract text from a DOCX file path using python-docx."""
    try:
        from docx import Document
        from docx.text.paragraph import Paragraph
        from docx.table import Table

        doc = Document(filepath)
        full_text = []
        for element in doc.element.body:
            if element.tag.endswith('p'):
                para = Paragraph(element, doc)
                if para.text.strip():
                    full_text.append(para.text)
            elif element.tag.endswith('tbl'):
                table = Table(element, doc)
                for row in table.rows:
                    # Deduplicate merged cells (same text repeated across merged columns)
                    seen = set()
                    row_text = []
                    for cell in row.cells:
                        txt = cell.text.strip()
                        if txt and txt not in seen:
                            seen.add(txt)
                            row_text.append(txt)
                    if row_text:
                        full_text.append(' | '.join(row_text))
        return '\n'.join(full_text)
    except ImportError:
        return "[DOCX extraction requires python-docx: pip install python-docx]"
    except Exception as e:  # noqa: BLE001  # broad catch: returns fallback
        return f"[Error extracting DOCX: {e}]"


def generate_worksheet_tool(title, worksheet_type, vocab_terms=None, questions=None,
                            summary_prompt=None, summary_key_points=None,
                            total_points=100, style_name=None, teacher_id='local-dev'):
    """Generate a .docx worksheet and save to Grading Setup."""
    require_teacher_id(teacher_id)
    try:
        from backend.services.worksheet_generator import generate_worksheet
        # Load subject from teacher settings
        settings = _load_settings(teacher_id)
        config = settings.get('config', {})
        subject = config.get('subject', '')
        return generate_worksheet(
            title=title,
            worksheet_type=worksheet_type,
            vocab_terms=vocab_terms,
            questions=questions,
            summary_prompt=summary_prompt,
            summary_key_points=summary_key_points,
            total_points=total_points,
            subject=subject,
            style_name=style_name
        )
    except ImportError:
        return {"error": "python-docx not installed. Run: pip install python-docx"}
    except Exception as e:  # noqa: BLE001  # broad catch: returns fallback
        return {"error": "Failed to generate worksheet: " + str(e)}


def generate_document_tool(title, content, style_name=None, save_to_builder=False, teacher_id='local-dev'):
    """Generate a formatted Word document with rich typography."""
    require_teacher_id(teacher_id)
    try:
        from backend.services.document_generator import generate_document
        return generate_document(
            title=title, content=content,
            style_name=style_name, save_to_builder=save_to_builder
        )
    except ImportError:
        return {"error": "python-docx not installed. Run: pip install python-docx"}
    except Exception as e:  # noqa: BLE001  # broad catch: returns fallback
        return {"error": "Failed to generate document: " + str(e)}


def generate_csv_tool(filename, headers, rows, teacher_id='local-dev'):
    """Generate a downloadable CSV or XLSX file based on filename extension."""
    require_teacher_id(teacher_id)
    from urllib.parse import quote

    EXPORT_DIR = graider_export_dir("Exports")
    os.makedirs(EXPORT_DIR, exist_ok=True)

    # Sanitize filename
    safe_name = "".join(c for c in filename if c.isalnum() or c in ' -_.').strip()
    is_xlsx = safe_name.lower().endswith('.xlsx')

    # Default to .xlsx if no recognized extension
    if not safe_name.lower().endswith(('.csv', '.xlsx')):
        safe_name += '.xlsx'
        is_xlsx = True

    filepath = os.path.join(EXPORT_DIR, safe_name)

    if is_xlsx:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quiz"

        # Write header row with styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Auto-fit column widths (approximate)
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in rows:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

        wb.save(filepath)
    else:
        import csv
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)

    download_url = "/api/download-csv/" + quote(safe_name)

    return {
        "status": "created",
        "filename": safe_name,
        "filepath": filepath,
        "download_url": download_url,
        "row_count": len(rows),
        "columns": headers,
    }


def save_document_style_tool(name, style, teacher_id='local-dev'):
    """Save a named visual style for documents."""
    require_teacher_id(teacher_id)
    try:
        from backend.services.document_generator import save_style
        return save_style(name=name, style_dict=style)
    except Exception as e:  # noqa: BLE001  # broad catch: returns fallback
        return {"error": "Failed to save style: " + str(e)}


def list_document_styles_tool(teacher_id='local-dev'):
    """List saved document visual styles."""
    require_teacher_id(teacher_id)
    try:
        from backend.services.document_generator import list_styles
        return list_styles()
    except Exception as e:  # noqa: BLE001  # broad catch: returns fallback
        return {"error": "Failed to list styles: " + str(e)}


def list_resources_tool(teacher_id='local-dev'):
    """List all uploaded supporting documents from the documents directory."""
    require_teacher_id(teacher_id)
    if not os.path.isdir(DOCUMENTS_DIR):
        return {"documents": [], "message": "No documents directory found. Upload documents in Settings > Resources."}

    documents = []
    try:
        for fname in sorted(os.listdir(DOCUMENTS_DIR)):
            if fname.endswith('.meta.json'):
                continue
            fpath = os.path.join(DOCUMENTS_DIR, fname)
            if not os.path.isfile(fpath):
                continue

            # Try to load metadata
            meta_path = fpath + ".meta.json"
            meta = {}
            if os.path.exists(meta_path):
                try:
                    with open(meta_path, 'r', encoding='utf-8') as f:
                        meta = json.load(f)
                except Exception:  # noqa: BLE001  # broad catch: error is logged
                    _logger.debug("document metadata load failed", exc_info=True)

            size_kb = round(os.path.getsize(fpath) / 1024, 1)
            documents.append({
                "filename": fname,
                "doc_type": meta.get("doc_type", "unknown"),
                "description": meta.get("description", ""),
                "size_kb": size_kb,
            })
    except Exception as e:  # noqa: BLE001  # broad catch: returns fallback
        return {"error": f"Error reading documents directory: {e}"}

    return {"documents": documents, "total": len(documents)}


def read_resource_tool(filename, teacher_id='local-dev'):
    """Read and return the text content of an uploaded document."""
    require_teacher_id(teacher_id)
    if not filename or not filename.strip():
        return {"error": "No filename provided"}

    # Prevent path traversal
    safe_name = os.path.basename(filename)
    filepath = os.path.join(DOCUMENTS_DIR, safe_name)

    if not os.path.exists(filepath):
        # Also check project root for built-in docs like User_Manual.md.
        # Use the shared PROJECT_ROOT, NOT a __file__-relative dirname chain: this file
        # moved one directory deeper in the package split, so the original 3-hop chain
        # would now resolve to backend/ instead of the repo root (behavior bug caught by
        # the Codex review of the split PR).
        project_root = PROJECT_ROOT
        alt_path = os.path.join(project_root, safe_name)
        if os.path.exists(alt_path):
            filepath = alt_path
        else:
            return {"error": f"Document not found: {safe_name}. Use list_resources to see available files."}

    ext = os.path.splitext(safe_name)[1].lower()
    pages = None

    try:
        if ext == '.pdf':
            content, pages = _extract_pdf_text(filepath)
        elif ext in ('.docx', '.doc'):
            content = _extract_docx_text(filepath)
        elif ext in ('.txt', '.md'):
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
        else:
            return {"error": f"Unsupported file type: {ext}. Supported: PDF, DOCX, DOC, TXT, MD."}
    except Exception as e:  # noqa: BLE001  # broad catch: returns fallback
        return {"error": f"Error reading {safe_name}: {e}"}

    truncated = False
    if len(content) > MAX_RESOURCE_TEXT:
        content = content[:MAX_RESOURCE_TEXT]
        truncated = True

    result = {
        "filename": safe_name,
        "content": content,
    }
    if pages:
        result["pages"] = pages
    if truncated:
        result["warning"] = f"Content truncated to {MAX_RESOURCE_TEXT} characters. Full document is larger."

    # Include metadata if available
    meta_path = filepath + ".meta.json"
    if os.path.exists(meta_path):
        try:
            with open(meta_path, 'r', encoding='utf-8') as f:
                meta = json.load(f)
            result["doc_type"] = meta.get("doc_type", "unknown")
            result["description"] = meta.get("description", "")
        except Exception:  # noqa: BLE001  # broad catch: error is logged
            _logger.debug("resource doc metadata enrich failed", exc_info=True)

    return result
