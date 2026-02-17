"""
Assistant Tools Service
=======================
Tool execution functions for the AI assistant.
Provides data querying, analytics, and Focus SIS automation.
"""

import os
import csv
import json
import subprocess
import statistics
from collections import defaultdict
from datetime import datetime


# Paths
RESULTS_FILE = os.path.expanduser("~/.graider_results.json")
SETTINGS_FILE = os.path.expanduser("~/.graider_global_settings.json")
ASSIGNMENTS_DIR = os.path.expanduser("~/.graider_assignments")
STUDENT_HISTORY_DIR = os.path.expanduser("~/.graider_data/student_history")
EXPORTS_DIR = os.path.expanduser("~/.graider_exports/focus")
CREDS_FILE = os.path.expanduser("~/.graider_data/portal_credentials.json")
LESSONS_DIR = os.path.expanduser("~/.graider_lessons")
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
STANDARDS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
PERIODS_DIR = os.path.expanduser("~/.graider_data/periods")
ACCOMMODATIONS_DIR = os.path.expanduser("~/.graider_data/accommodations")
PARENT_CONTACTS_FILE = os.path.expanduser("~/.graider_data/parent_contacts.json")
PERIOD_CSVS_DIR = os.path.join(PROJECT_ROOT, "Period CSVs")
MEMORY_FILE = os.path.expanduser("~/.graider_data/assistant_memory.json")
DOCUMENTS_DIR = os.path.expanduser("~/.graider_data/documents")
MAX_MEMORIES = 50


def _fuzzy_name_match(search, full_name):
    """Word-based name matching. Returns True if every word in search appears
    as a word (or word-prefix) in full_name. Order-independent, case-insensitive.

    Examples:
        _fuzzy_name_match("Dicen Wilkins", "Dicen Macheil Wilkins Reels") → True
        _fuzzy_name_match("Dicen Wilkins", "Wilkins Reels, Dicen Macheil") → True
        _fuzzy_name_match("Luke Lundell", "Luke J Lundell") → True
        _fuzzy_name_match("John Smith", "Jane Smith") → False
    """
    import re
    # Strip punctuation (commas, semicolons, periods) and normalize
    clean = lambda s: re.sub(r'[,;.\'"]+', ' ', s.lower()).split()
    search_words = clean(search)
    name_words = clean(full_name)
    if not search_words:
        return False
    return all(
        any(nw.startswith(sw) or sw.startswith(nw) for nw in name_words)
        for sw in search_words
    )


def _extract_first_name(name):
    """Extract the actual first name from various name formats.

    Handles:
        "First Last"                → "First"
        "Last, First Middle"        → "First"
        "Last; First Middle"        → "First"
        "First Middle Last Last2"   → "First"
    """
    if not name or name == 'Student':
        return 'Student'
    name = name.strip()
    # If comma or semicolon present, part after separator is first name
    for sep in [',', ';']:
        if sep in name:
            parts = name.split(sep, 1)
            after = parts[1].strip()
            return after.split()[0] if after else parts[0].strip().split()[0]
    return name.split()[0]


def _safe_int_score(val):
    """Safely convert a score value to int (handles str, float, None)."""
    try:
        return int(float(val)) if val else 0
    except (ValueError, TypeError):
        return 0


def _normalize_period(period):
    """Normalize period name: 'Period_5' -> 'Period 5', 'period 5' -> 'Period 5', '5' -> 'Period 5'."""
    import re
    p = period.strip().replace('_', ' ')
    # Bare number: '2' -> 'Period 2'
    if re.match(r'^\d+$', p):
        return f'Period {p}'
    # Capitalize: 'period 5' -> 'Period 5'
    if p.lower().startswith('period'):
        return 'Period' + p[6:]
    return p


def _normalize_assignment_name(name):
    """Normalize assignment name for comparison (strips suffixes like (1), .docx)."""
    import re
    n = name.strip()
    n = re.sub(r'\s*\(\d+\)\s*$', '', n)       # Remove trailing (1), (2)
    n = re.sub(r'\.docx?\s*$', '', n, flags=re.IGNORECASE)  # Remove .docx
    n = re.sub(r'\.pdf\s*$', '', n, flags=re.IGNORECASE)    # Remove .pdf
    return n.strip()


def _get_period_assignments(rows):
    """Build a map of normalized_period -> set of normalized assignment names.
    Also returns the display name for each normalized assignment.
    Merges truncated assignment names with their full versions."""
    from collections import defaultdict
    period_assigns = defaultdict(set)
    assign_display = {}  # normalized -> best display name
    for row in rows:
        period = _normalize_period(row.get("period", "") or row.get("quarter", ""))
        assign = row.get("assignment", "")
        if period and assign:
            norm = _normalize_assignment_name(assign)
            period_assigns[period].add(norm)
            # Keep the longest display name (most descriptive)
            if norm not in assign_display or len(assign) > len(assign_display[norm]):
                assign_display[norm] = assign

    # Merge truncated names: if one norm is a prefix of another (>=20 chars), unify them
    all_norms = sorted(assign_display.keys(), key=len, reverse=True)
    merge_map = {}  # short_norm -> long_norm
    for i, short in enumerate(all_norms):
        if short in merge_map:
            continue
        short_lower = short.lower()
        for long in all_norms[:i]:
            if long in merge_map:
                continue
            if len(short_lower) >= 20 and long.lower().startswith(short_lower) and short_lower != long.lower():
                merge_map[short] = long
                break

    if merge_map:
        for period in period_assigns:
            merged = set()
            for n in period_assigns[period]:
                merged.add(merge_map.get(n, n))
            period_assigns[period] = merged
        for short, long in merge_map.items():
            if short in assign_display:
                # Keep the longer display name
                if long not in assign_display or len(assign_display[short]) > len(assign_display[long]):
                    assign_display[long] = assign_display[short]
                del assign_display[short]

    return period_assigns, assign_display, merge_map


def _get_output_folder():
    """Get the configured output folder for grading results."""
    output_folder = os.path.expanduser("~/Downloads/Graider/Results")
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r') as f:
                settings = json.load(f)
                output_folder = settings.get('output_folder', output_folder)
        except Exception:
            pass
    return output_folder


def _load_results():
    """Load grading results from the results JSON file."""
    if not os.path.exists(RESULTS_FILE):
        return []
    try:
        with open(RESULTS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []


def _load_master_csv(period_filter='all'):
    """Load and parse the master grades CSV, then merge in any results from
    the results JSON that aren't already present. This ensures the Assistant
    always sees the most complete, up-to-date data."""
    import re

    def _norm_assign(name):
        n = name.strip()
        n = re.sub(r'\s*\(\d+\)\s*$', '', n)
        n = re.sub(r'\.docx?\s*$', '', n, flags=re.IGNORECASE)
        n = re.sub(r'\.pdf\s*$', '', n, flags=re.IGNORECASE)
        return n.strip().lower()

    output_folder = _get_output_folder()
    master_file = os.path.join(output_folder, "master_grades.csv")

    rows = []
    seen_keys = set()  # (student_id, normalized_assignment)

    # 1. Load master CSV
    if os.path.exists(master_file):
        try:
            with open(master_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if not row.get("Student Name"):
                        continue
                    if period_filter != 'all' and row.get("Quarter", "") != period_filter:
                        continue
                    sid = row.get("Student ID", "")
                    assign = row.get("Assignment", "")
                    parsed = {
                        "student_name": row.get("Student Name", ""),
                        "student_id": sid,
                        "first_name": row.get("First Name", ""),
                        "date": row.get("Date", ""),
                        "assignment": assign,
                        "period": row.get("Period", ""),
                        "quarter": row.get("Quarter", ""),
                        "score": int(float(row.get("Overall Score", 0) or 0)),
                        "letter_grade": row.get("Letter Grade", ""),
                        "content": int(float(row.get("Content Accuracy", 0) or 0)),
                        "completeness": int(float(row.get("Completeness", 0) or 0)),
                        "writing": int(float(row.get("Writing Quality", 0) or 0)),
                        "effort": int(float(row.get("Effort Engagement", 0) or 0)),
                    }
                    rows.append(parsed)
                    if sid and sid != "UNKNOWN":
                        seen_keys.add((sid, _norm_assign(assign)))
        except Exception:
            pass

    # 2. Merge in results JSON entries not already in master CSV.
    #    Master CSV is authoritative for scores (it's synced on edits).
    #    Results JSON fills in grades that haven't been written to CSV yet.
    results_json = _load_results()

    # Build a name→student_id lookup from known data for resolving UNKNOWN IDs
    name_to_id = {}
    name_to_period = {}
    for row in rows:
        sid = row["student_id"]
        name = row["student_name"].lower().strip()
        if sid and sid != "UNKNOWN" and name:
            name_to_id[name] = sid
            if row["period"]:
                name_to_period[name] = row["period"]

    for r in results_json:
        sid = str(r.get("student_id", ""))
        rname = r.get("student_name", "")
        assign = r.get("assignment", "")
        if not assign or not rname:
            continue

        # Resolve UNKNOWN student_id by matching name against known students
        if not sid or sid == "UNKNOWN":
            rname_lower = rname.lower().strip()
            rname_words = rname_lower.split()
            resolved_sid = None
            for known_name, known_id in name_to_id.items():
                # Exact substring match
                if len(rname_lower) >= 5 and (rname_lower in known_name or known_name.startswith(rname_lower)):
                    resolved_sid = known_id
                    break
                # Word-prefix match: every word in the short name starts a word in the known name
                # Handles "vincent scar" → "vincent ray scarola"
                if len(rname_words) >= 2:
                    known_words = known_name.split()
                    if all(any(kw.startswith(rw) for kw in known_words) for rw in rname_words):
                        resolved_sid = known_id
                        break
            if resolved_sid:
                sid = resolved_sid
            else:
                continue  # Can't resolve — skip

        key = (sid, _norm_assign(assign))
        if key in seen_keys:
            # Already in master CSV — master CSV is authoritative, don't overwrite
            continue

        # Not in master CSV — add it from results JSON
        breakdown = r.get("breakdown", {})
        period = r.get("period", "") or name_to_period.get(rname.lower().strip(), "")
        if period_filter != 'all' and period_filter not in (period, ""):
            continue
        rows.append({
            "student_name": rname if len(rname) > 5 else next((row["student_name"] for row in rows if row["student_id"] == sid), rname),
            "student_id": sid,
            "first_name": _extract_first_name(rname),
            "date": r.get("graded_at", "")[:10] if r.get("graded_at") else "",
            "assignment": assign,
            "period": period,
            "quarter": "",
            "score": int(float(r.get("score", 0) or 0)),
            "letter_grade": r.get("letter_grade", ""),
            "content": int(float(breakdown.get("content_accuracy", 0) or 0)),
            "completeness": int(float(breakdown.get("completeness", 0) or 0)),
            "writing": int(float(breakdown.get("writing_quality", 0) or 0)),
            "effort": int(float(breakdown.get("effort_engagement", 0) or 0)),
        })
        seen_keys.add(key)

    return rows


# ═══════════════════════════════════════════════════════
# TOOL DEFINITIONS (for Anthropic tool use)
# ═══════════════════════════════════════════════════════

TOOL_DEFINITIONS = [
    {
        "name": "query_grades",
        "description": "Search and filter student grades. Use this when the teacher asks about specific students, assignments, score ranges, or periods.",
        "input_schema": {
            "type": "object",
            "properties": {
                "student_name": {
                    "type": "string",
                    "description": "Student name to search for (partial match, case-insensitive)"
                },
                "assignment": {
                    "type": "string",
                    "description": "Assignment name to filter by (partial match)"
                },
                "period": {
                    "type": "string",
                    "description": "Grading period/quarter to filter by"
                },
                "min_score": {
                    "type": "number",
                    "description": "Minimum score threshold"
                },
                "max_score": {
                    "type": "number",
                    "description": "Maximum score threshold"
                },
                "letter_grade": {
                    "type": "string",
                    "description": "Letter grade to filter by (A, B, C, D, F)"
                },
                "limit": {
                    "type": "integer",
                    "description": "Maximum number of results to return (default 25)"
                }
            }
        }
    },
    {
        "name": "get_student_summary",
        "description": "Get a comprehensive summary of a specific student's performance including all grades, average, trend, category breakdowns, and strengths/weaknesses.",
        "input_schema": {
            "type": "object",
            "properties": {
                "student_name": {
                    "type": "string",
                    "description": "Student name to look up (partial match)"
                }
            },
            "required": ["student_name"]
        }
    },
    {
        "name": "get_class_analytics",
        "description": "Get class-wide analytics including average, grade distribution, top/bottom performers, and students needing attention. Optionally filter by period.",
        "input_schema": {
            "type": "object",
            "properties": {
                "period": {
                    "type": "string",
                    "description": "Grading period/quarter to filter by (omit for all periods)"
                }
            }
        }
    },
    {
        "name": "get_assignment_stats",
        "description": "Get statistics for a specific assignment including count, mean, median, min, max, standard deviation, and grade distribution.",
        "input_schema": {
            "type": "object",
            "properties": {
                "assignment_name": {
                    "type": "string",
                    "description": "Assignment name to look up (partial match)"
                }
            },
            "required": ["assignment_name"]
        }
    },
    {
        "name": "list_assignments",
        "description": "List all graded assignments with their student count and average score.",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "create_focus_assignment",
        "description": "Create an assignment in Focus gradebook via browser automation. Requires VPortal credentials to be configured in Settings. The browser will open for the teacher to complete 2FA and verify before saving.",
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {
                    "type": "string",
                    "description": "Assignment name"
                },
                "category": {
                    "type": "string",
                    "description": "Category (e.g., 'Assessments', 'Classwork')"
                },
                "points": {
                    "type": "integer",
                    "description": "Point value for the assignment"
                },
                "date": {
                    "type": "string",
                    "description": "Due date in MM/DD/YYYY format"
                },
                "description": {
                    "type": "string",
                    "description": "Assignment description"
                }
            },
            "required": ["name"]
        }
    },
    {
        "name": "export_grades_csv",
        "description": "Export grades as a Focus SIS-compatible CSV file. Generates per-period CSVs with Student ID and Score columns.",
        "input_schema": {
            "type": "object",
            "properties": {
                "assignment": {
                    "type": "string",
                    "description": "Assignment name to filter export (omit for all)"
                },
                "period": {
                    "type": "string",
                    "description": "Period to filter export (omit for all periods)"
                }
            }
        }
    },
    {
        "name": "analyze_grade_causes",
        "description": "Deep analysis of WHY students got the grades they did on an assignment. Shows rubric category breakdown averages, most commonly missed/unanswered questions, score distribution by category, and identifies which sections or skills caused the most point loss. Use this when the teacher asks about causes of low grades, common mistakes, or what students struggled with.",
        "input_schema": {
            "type": "object",
            "properties": {
                "assignment_name": {
                    "type": "string",
                    "description": "Assignment name (partial match)"
                },
                "period": {
                    "type": "string",
                    "description": "Filter by period (optional)"
                },
                "score_threshold": {
                    "type": "number",
                    "description": "Only analyze students below this score (optional, e.g. 70 for failing students)"
                }
            },
            "required": ["assignment_name"]
        }
    },
    {
        "name": "get_feedback_patterns",
        "description": "Analyze feedback text across an assignment to find common themes, recurring issues, and patterns in what students did well or poorly. Extracts strengths and developing skills aggregated across all students. Use when teacher asks about common mistakes, patterns, or what feedback was given.",
        "input_schema": {
            "type": "object",
            "properties": {
                "assignment_name": {
                    "type": "string",
                    "description": "Assignment name (partial match)"
                },
                "period": {
                    "type": "string",
                    "description": "Filter by period (optional)"
                }
            },
            "required": ["assignment_name"]
        }
    },
    {
        "name": "compare_periods",
        "description": "Compare performance across class periods for a specific assignment or overall. Shows average scores, grade distributions, and category breakdowns per period. Use when teacher asks how different classes compared or which period did best/worst.",
        "input_schema": {
            "type": "object",
            "properties": {
                "assignment_name": {
                    "type": "string",
                    "description": "Assignment name to compare across periods (optional, omit for overall)"
                }
            }
        }
    },
    {
        "name": "lookup_student_info",
        "description": "Look up student contact and roster information. Returns student ID, local ID, grade level, period, course codes, student email, parent emails, parent phone numbers, 504 plan status, detailed contacts (up to 3 guardians with names, relationships, roles), and full student schedule (all periods with teachers and courses). Search by student name, ID, or period. Supports BATCH lookup via student_ids array — use this to get parent contacts for multiple students in one call (e.g., after querying grades to find failing students).",
        "input_schema": {
            "type": "object",
            "properties": {
                "student_name": {
                    "type": "string",
                    "description": "Student name to search for (partial match, case-insensitive)"
                },
                "student_id": {
                    "type": "string",
                    "description": "Single student ID number to look up directly"
                },
                "student_ids": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of student ID numbers for batch lookup. Use this to get contacts for multiple students at once."
                },
                "period": {
                    "type": "string",
                    "description": "List all students in this period (e.g., 'Period 1', '1')"
                }
            }
        }
    },
    {
        "name": "get_missing_assignments",
        "description": "Find missing/unsubmitted assignments. Three modes: (1) by student name — shows which assignments they submitted and which are missing, (2) by period — shows all students with missing work, (3) by assignment — shows which students haven't submitted a specific assignment. Compares each student's grades against all assignments graded in their period.",
        "input_schema": {
            "type": "object",
            "properties": {
                "student_name": {
                    "type": "string",
                    "description": "Student name to check missing work for (partial match)"
                },
                "period": {
                    "type": "string",
                    "description": "Period to check (e.g., 'Period 2', '2'). Shows all students with missing work."
                },
                "assignment_name": {
                    "type": "string",
                    "description": "Assignment name to check (partial match). Shows which students haven't submitted it."
                }
            }
        }
    },
    {
        "name": "generate_worksheet",
        "description": "Generate a structured worksheet document (Cornell Notes, Fill-in-the-Blank, short-answer, vocabulary) from a reading or topic. Creates a downloadable Word document with an embedded invisible answer key for consistent AI grading. Automatically saved to Grading Setup. Use when the teacher asks to create a worksheet, assignment, or activity from a reading, textbook page, or topic.",
        "input_schema": {
            "type": "object",
            "properties": {
                "title": {
                    "type": "string",
                    "description": "Worksheet title (e.g., 'Cornell Notes - Expanding into Native American Lands')"
                },
                "worksheet_type": {
                    "type": "string",
                    "enum": ["cornell-notes", "fill-in-blank", "short-answer", "vocabulary"],
                    "description": "Type of worksheet to generate"
                },
                "vocab_terms": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "term": {"type": "string"},
                            "definition": {"type": "string", "description": "Expected definition (for answer key)"}
                        },
                        "required": ["term"]
                    },
                    "description": "Vocabulary terms with expected definitions"
                },
                "questions": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "question": {"type": "string"},
                            "expected_answer": {"type": "string", "description": "Expected answer (for answer key)"},
                            "points": {"type": "integer", "default": 10}
                        },
                        "required": ["question"]
                    },
                    "description": "Questions with expected answers"
                },
                "summary_prompt": {
                    "type": "string",
                    "description": "Instruction for the summary section (e.g., 'Summarize the reading in 3-5 sentences...')"
                },
                "summary_key_points": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Key points that should appear in a good summary"
                },
                "total_points": {
                    "type": "integer",
                    "default": 100,
                    "description": "Total point value for the worksheet"
                },
                "style_name": {
                    "type": "string",
                    "description": "Name of a saved visual style to apply. Omit to use defaults."
                }
            },
            "required": ["title", "worksheet_type"]
        }
    },
    {
        "name": "generate_document",
        "description": "Generate a formatted Word document (.docx) with rich typography. Use for study guides, reference sheets, parent letters, lesson outlines, rubrics, or any document the teacher requests. NOT for gradeable worksheets (use generate_worksheet for those).",
        "input_schema": {
            "type": "object",
            "properties": {
                "title": {
                    "type": "string",
                    "description": "Document title (main heading)"
                },
                "content": {
                    "type": "array",
                    "description": "Ordered content blocks. Text supports **bold** and *italic* markdown.",
                    "items": {
                        "type": "object",
                        "properties": {
                            "type": {"type": "string", "enum": ["heading", "paragraph", "bullet_list", "numbered_list", "table"]},
                            "text": {"type": "string", "description": "Text content (heading/paragraph)"},
                            "level": {"type": "integer", "description": "Heading level 1-3"},
                            "items": {"type": "array", "items": {"type": "string"}, "description": "List items"},
                            "rows": {"type": "array", "items": {"type": "array", "items": {"type": "string"}}, "description": "Table rows (first = header)"}
                        },
                        "required": ["type"]
                    }
                },
                "style_name": {
                    "type": "string",
                    "description": "Name of a saved visual style to apply. Omit to use defaults."
                },
                "save_to_builder": {
                    "type": "boolean",
                    "description": "If true, also save this document to Grading Setup for grading. Only set true if the teacher confirms they want it saved."
                }
            },
            "required": ["title", "content"]
        }
    },
    {
        "name": "save_document_style",
        "description": "Save a visual style (fonts, sizes, colors, spacing) so future documents of this type always look the same. Use when the teacher says they like how a document looks and want to reuse that formatting.",
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Style name (e.g., 'cornell-notes', 'parent-letter')"},
                "style": {
                    "type": "object",
                    "description": "Visual properties to save",
                    "properties": {
                        "title_font_name": {"type": "string"},
                        "title_font_size": {"type": "integer"},
                        "title_bold": {"type": "boolean"},
                        "heading_font_name": {"type": "string"},
                        "heading_sizes": {"type": "object"},
                        "body_font_name": {"type": "string"},
                        "body_font_size": {"type": "integer"},
                        "line_spacing": {"type": "number"},
                        "table_header_bg": {"type": "string"},
                        "accent_color": {"type": "string"}
                    }
                }
            },
            "required": ["name", "style"]
        }
    },
    {
        "name": "list_document_styles",
        "description": "List saved document visual styles. Use before generating a document to check if a matching style exists.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "generate_csv",
        "description": "Generate a downloadable spreadsheet file (XLSX or CSV). Use .xlsx for Wayground quizzes (required format) and any polished spreadsheet. Use .csv for simple data exports. When generating a Wayground quiz, use .xlsx and match the EXACT column structure from the Assessment Templates in your context (Question Text, Question Type, Option 1-5, Correct Answer, Time in seconds, Image Link, Answer explanation).",
        "input_schema": {
            "type": "object",
            "properties": {
                "filename": {
                    "type": "string",
                    "description": "Name for the file. Use .xlsx for Wayground and polished spreadsheets, .csv for simple exports. Defaults to .xlsx if no extension given."
                },
                "headers": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Column headers for the CSV"
                },
                "rows": {
                    "type": "array",
                    "items": {
                        "type": "array",
                        "items": {"type": "string"}
                    },
                    "description": "Data rows. Each row is an array of string values matching the headers."
                }
            },
            "required": ["filename", "headers", "rows"]
        }
    },
    {
        "name": "recommend_next_lesson",
        "description": "Analyze student performance and recommend what the next lesson should focus on. Provides DIFFERENTIATED recommendations by class level (advanced/standard/support periods) with DOK-appropriate standards. Also identifies IEP/504 accommodation patterns — whether accommodated students struggled differently and what modifications may help. Cross-references rubric breakdowns, unanswered questions, developing skills, and curriculum standards. Use when teacher asks 'what should I teach next?', 'what do students need to work on?', or 'how should I differentiate the next lesson?'.",
        "input_schema": {
            "type": "object",
            "properties": {
                "assignment_name": {
                    "type": "string",
                    "description": "Assignment to base recommendations on (partial match). Omit to use the most recent assignment."
                },
                "period": {
                    "type": "string",
                    "description": "Focus on a specific period (optional). Different periods may have different class levels and weaknesses."
                },
                "num_assignments": {
                    "type": "integer",
                    "description": "Number of recent assignments to analyze (default 1, max 5). More assignments = broader trend analysis."
                }
            }
        }
    },
    {
        "name": "save_memory",
        "description": "Save an important fact about the teacher or their classes for future conversations. Use this when the teacher shares preferences, class structure, workflow habits, or any information that would be useful to remember across sessions. Only save genuinely useful facts, not temporary or conversation-specific details.",
        "input_schema": {
            "type": "object",
            "properties": {
                "fact": {
                    "type": "string",
                    "description": "The fact to remember (e.g., 'Period 3 is honors with higher expectations')"
                }
            },
            "required": ["fact"]
        }
    },
    {
        "name": "get_standards",
        "description": "Look up curriculum standards for the teacher's state and subject. Returns ALL standards when no topic is specified, or filter by keyword. Use this to get full details (vocabulary, learning targets, essential questions) for standards. For a quick overview of all standards, use list_all_standards first.",
        "input_schema": {
            "type": "object",
            "properties": {
                "topic": {
                    "type": "string",
                    "description": "Topic keyword to filter standards (e.g., 'fractions', 'civil war', 'photosynthesis'). Matches against benchmark text, topics, and vocabulary."
                },
                "dok_max": {
                    "type": "integer",
                    "description": "Maximum DOK level to include (1-4). Use 2 for support classes, 3 for standard, 4 for advanced."
                }
            }
        }
    },
    {
        "name": "get_recent_lessons",
        "description": "List saved lesson plans, optionally filtered by unit name. Shows what has been taught recently — topics, standards covered, vocabulary, and objectives per day. Use this when the teacher asks to create a quiz or worksheet 'for this unit', 'for what we've been doing', or references past lessons.",
        "input_schema": {
            "type": "object",
            "properties": {
                "unit_name": {
                    "type": "string",
                    "description": "Filter by unit name (partial match, case-insensitive). Omit to show all units and recent lessons."
                }
            }
        }
    },
    {
        "name": "get_calendar",
        "description": "Read the teaching calendar — scheduled lessons and holidays for a date range. AUTHORITATIVE source for what the teacher is teaching. If scheduled_lessons is non-empty, those lessons ARE scheduled — always reference them by title and topic. Use when the teacher asks 'what am I teaching this week?', 'what's on my calendar?', 'generate a worksheet for Tuesday', or anything about their schedule. When asked about a specific day, set both start_date and end_date to that exact date.",
        "input_schema": {
            "type": "object",
            "properties": {
                "start_date": {
                    "type": "string",
                    "description": "Start date in YYYY-MM-DD format. Defaults to today. For a specific day like 'Tuesday Feb 17', use '2026-02-17'."
                },
                "end_date": {
                    "type": "string",
                    "description": "End date in YYYY-MM-DD format. Defaults to 7 days from start. For a specific day, set equal to start_date."
                }
            }
        }
    },
    {
        "name": "schedule_lesson",
        "description": "Schedule a saved lesson plan onto the teaching calendar on a specific date. Use when the teacher says 'schedule the Revolution unit starting Monday', 'put Unit 3 on the calendar', or asks to plan out their week/month. For multi-day lessons, call this once per day with incrementing day_number and dates.",
        "input_schema": {
            "type": "object",
            "properties": {
                "date": {
                    "type": "string",
                    "description": "Date to schedule in YYYY-MM-DD format"
                },
                "unit": {
                    "type": "string",
                    "description": "Unit name the lesson belongs to"
                },
                "lesson_title": {
                    "type": "string",
                    "description": "Title of the lesson"
                },
                "day_number": {
                    "type": "integer",
                    "description": "Day number within the lesson (e.g., 1 for Day 1). Optional."
                },
                "lesson_file": {
                    "type": "string",
                    "description": "Path to lesson file like 'Unit Name/Lesson Title.json'. Optional."
                }
            },
            "required": ["date", "lesson_title"]
        }
    },
    {
        "name": "add_calendar_holiday",
        "description": "Add a holiday or break to the teaching calendar. Use when the teacher says 'add Spring Break', 'mark Monday as a holiday', or 'we're off next Friday'.",
        "input_schema": {
            "type": "object",
            "properties": {
                "date": {
                    "type": "string",
                    "description": "Start date in YYYY-MM-DD format"
                },
                "name": {
                    "type": "string",
                    "description": "Name of the holiday or break (e.g., 'Spring Break', 'Teacher Workday')"
                },
                "end_date": {
                    "type": "string",
                    "description": "End date for multi-day breaks in YYYY-MM-DD format. Omit for single-day holidays."
                }
            },
            "required": ["date", "name"]
        }
    },
    {
        "name": "list_all_standards",
        "description": "Get a compact index of ALL curriculum standards for the teacher's subject. Returns every standard code with a short benchmark summary. Use this first to see what standards exist, then use get_standards with a topic keyword to get full details (vocabulary, learning targets, essential questions) for specific standards.",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "list_resources",
        "description": "List all supporting documents uploaded by the teacher (curriculum guides, pacing calendars, rubrics, etc.). Shows filename, type, description, and size. Use this to discover what reference materials are available before reading specific ones with read_resource.",
        "input_schema": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "read_resource",
        "description": "Read the full text content of an uploaded supporting document. Use after list_resources to read a specific file. Supports PDF, DOCX, DOC, TXT, and MD files. Use this when the teacher asks about pacing, curriculum content, or when you need reference material for generating lessons, worksheets, or assessments.",
        "input_schema": {
            "type": "object",
            "properties": {
                "filename": {
                    "type": "string",
                    "description": "Exact filename of the document to read (from list_resources results)"
                }
            },
            "required": ["filename"]
        }
    }
]


# ═══════════════════════════════════════════════════════
# TOOL EXECUTION FUNCTIONS
# ═══════════════════════════════════════════════════════

def query_grades(student_name=None, assignment=None, period=None,
                 min_score=None, max_score=None, letter_grade=None, limit=25):
    """Search and filter student grades."""
    rows = _load_master_csv(period_filter=period or 'all')
    results_json = _load_results()

    # Build a lookup from results JSON for feedback
    feedback_lookup = {}
    for r in results_json:
        key = (r.get('student_name', ''), r.get('assignment', ''))
        feedback_lookup[key] = r.get('feedback', '')

    filtered = []
    for row in rows:
        if student_name:
            if student_name.lower() not in row["student_name"].lower():
                continue
        if assignment:
            if assignment.lower() not in row["assignment"].lower():
                continue
        if min_score is not None and row["score"] < min_score:
            continue
        if max_score is not None and row["score"] > max_score:
            continue
        if letter_grade and row["letter_grade"].upper() != letter_grade.upper():
            continue

        # Get feedback if available (truncated)
        fb = feedback_lookup.get((row["student_name"], row["assignment"]), "")
        if len(fb) > 150:
            fb = fb[:150] + "..."

        filtered.append({
            "student_name": row["student_name"],
            "student_id": row.get("student_id", ""),
            "assignment": row["assignment"],
            "score": row["score"],
            "letter_grade": row["letter_grade"],
            "period": row["quarter"],
            "date": row["date"],
            "feedback_preview": fb
        })

    total = len(filtered)
    filtered = filtered[:limit]

    return {
        "results": filtered,
        "total_matches": total,
        "showing": len(filtered)
    }


def get_student_summary(student_name):
    """Get comprehensive summary for a specific student."""
    rows = _load_master_csv()

    # Find matching student (fuzzy word match — handles compound names)
    student_rows = [r for r in rows if _fuzzy_name_match(student_name, r["student_name"])]

    if not student_rows:
        return {"error": f"No student found matching '{student_name}'"}

    # Get the actual full name and student ID from the first match
    actual_name = student_rows[0]["student_name"]
    actual_id = student_rows[0].get("student_id", "")
    # Re-filter: match by student_id (preferred) or exact name
    # This catches rows with truncated names but the same student_id
    if actual_id and actual_id != "UNKNOWN":
        student_rows = [r for r in rows if r.get("student_id") == actual_id]
        # Use the longest name variant as the display name
        for r in student_rows:
            if len(r["student_name"]) > len(actual_name):
                actual_name = r["student_name"]
    else:
        student_rows = [r for r in rows if r["student_name"] == actual_name]

    sorted_rows = sorted(student_rows, key=lambda x: x["date"])
    scores = [r["score"] for r in student_rows]
    avg = round(sum(scores) / len(scores), 1) if scores else 0

    # Trend calculation
    if len(sorted_rows) >= 2:
        first_half = sorted_rows[:len(sorted_rows) // 2]
        second_half = sorted_rows[len(sorted_rows) // 2:]
        first_avg = sum(r["score"] for r in first_half) / len(first_half)
        second_avg = sum(r["score"] for r in second_half) / len(second_half)
        diff = second_avg - first_avg
        if diff > 3:
            trend = "improving"
        elif diff < -3:
            trend = "declining"
        else:
            trend = "stable"
    else:
        trend = "insufficient data"

    # Category averages
    cat_scores = {"content": [], "completeness": [], "writing": [], "effort": []}
    for r in student_rows:
        cat_scores["content"].append(r["content"])
        cat_scores["completeness"].append(r["completeness"])
        cat_scores["writing"].append(r["writing"])
        cat_scores["effort"].append(r["effort"])

    cat_avgs = {}
    for cat, vals in cat_scores.items():
        cat_avgs[cat] = round(sum(vals) / len(vals), 1) if vals else 0

    # Determine strengths/weaknesses
    sorted_cats = sorted(cat_avgs.items(), key=lambda x: x[1], reverse=True)
    strengths = [c[0] for c in sorted_cats[:2] if c[1] > 0]
    weaknesses = [c[0] for c in sorted_cats[-2:] if c[1] > 0]

    # Grade history
    grade_history = [
        {"assignment": r["assignment"], "score": r["score"],
         "letter_grade": r["letter_grade"], "date": r["date"]}
        for r in sorted_rows
    ]

    # Determine student's period and find missing assignments
    student_period = ""
    for r in student_rows:
        p = _normalize_period(r.get("period", "") or r.get("quarter", ""))
        if p:
            student_period = p
            break

    missing_assignments = []
    all_period_assigns = set()
    if student_period:
        period_assigns, assign_display, merge_map = _get_period_assignments(rows)
        all_period_assigns = period_assigns.get(student_period, set())
        student_has = set(merge_map.get(_normalize_assignment_name(r["assignment"]), _normalize_assignment_name(r["assignment"])) for r in student_rows)
        missing_norms = all_period_assigns - student_has
        missing_assignments = sorted(assign_display.get(n, n) for n in missing_norms)

    return {
        "student_name": actual_name,
        "period": student_period,
        "total_assignments_graded": len(student_rows),
        "total_assignments_in_period": len(all_period_assigns) if student_period else 0,
        "missing_assignments": missing_assignments,
        "missing_count": len(missing_assignments),
        "average_score": avg,
        "trend": trend,
        "highest_score": max(scores) if scores else 0,
        "lowest_score": min(scores) if scores else 0,
        "category_averages": cat_avgs,
        "strengths": strengths,
        "weaknesses": weaknesses,
        "grade_history": grade_history
    }


def get_class_analytics(period=None):
    """Get class-wide analytics."""
    rows = _load_master_csv(period_filter=period or 'all')

    if not rows:
        return {"error": "No grading data available"}

    students = defaultdict(list)
    assignments = defaultdict(list)

    for row in rows:
        students[row["student_name"]].append(row)
        assignments[row["assignment"]].append(row)

    all_scores = [r["score"] for r in rows]
    class_avg = round(sum(all_scores) / len(all_scores), 1) if all_scores else 0

    # Get roster count for the requested period (authoritative student count)
    roster = _load_roster()
    if period and period != 'all':
        roster_for_period = [s for s in roster if s["period"].lower() == period.lower()]
    else:
        roster_for_period = roster
    roster_student_count = len(roster_for_period) if roster_for_period else None

    # Grade distribution by unique student averages, not raw submissions
    student_avg_scores = []
    for name, grades in students.items():
        s_scores = [g["score"] for g in grades]
        student_avg_scores.append(round(sum(s_scores) / len(s_scores), 1))

    grade_dist = {
        "A": len([s for s in student_avg_scores if s >= 90]),
        "B": len([s for s in student_avg_scores if 80 <= s < 90]),
        "C": len([s for s in student_avg_scores if 70 <= s < 80]),
        "D": len([s for s in student_avg_scores if 60 <= s < 70]),
        "F": len([s for s in student_avg_scores if s < 60])
    }

    # Student averages for ranking
    student_avgs = []
    for name, grades in students.items():
        s_scores = [g["score"] for g in grades]
        avg = round(sum(s_scores) / len(s_scores), 1)
        sorted_grades = sorted(grades, key=lambda x: x["date"])
        if len(sorted_grades) >= 2:
            trend = "improving" if sorted_grades[-1]["score"] > sorted_grades[0]["score"] \
                else "declining" if sorted_grades[-1]["score"] < sorted_grades[0]["score"] \
                else "stable"
        else:
            trend = "stable"
        student_avgs.append({"name": name, "average": avg, "trend": trend,
                             "assignments_count": len(grades)})

    student_avgs.sort(key=lambda x: x["average"], reverse=True)
    top_performers = student_avgs[:5]
    attention_needed = [s for s in student_avgs if s["average"] < 70 or s["trend"] == "declining"]

    # Use roster count when available, fall back to graded student count
    total_students = roster_student_count if roster_student_count else len(students)
    students_not_graded = max(0, total_students - len(students)) if roster_student_count else 0

    return {
        "class_average": class_avg,
        "total_students": total_students,
        "students_with_grades": len(students),
        "students_not_graded": students_not_graded,
        "total_grades": len(rows),
        "grade_distribution": grade_dist,
        "top_performers": top_performers,
        "attention_needed": attention_needed,
        "period": period or "all"
    }


def get_assignment_stats(assignment_name):
    """Get statistics for a specific assignment."""
    rows = _load_master_csv()

    # Partial match on assignment name
    matched = [r for r in rows if assignment_name.lower() in r["assignment"].lower()]

    if not matched:
        return {"error": f"No assignment found matching '{assignment_name}'"}

    # Get exact assignment name from first match
    actual_name = matched[0]["assignment"]
    matched = [r for r in rows if r["assignment"] == actual_name]

    scores = [r["score"] for r in matched]

    grade_dist = {
        "A": len([s for s in scores if s >= 90]),
        "B": len([s for s in scores if 80 <= s < 90]),
        "C": len([s for s in scores if 70 <= s < 80]),
        "D": len([s for s in scores if 60 <= s < 70]),
        "F": len([s for s in scores if s < 60])
    }

    return {
        "assignment_name": actual_name,
        "count": len(scores),
        "mean": round(sum(scores) / len(scores), 1) if scores else 0,
        "median": round(statistics.median(scores), 1) if scores else 0,
        "min": min(scores) if scores else 0,
        "max": max(scores) if scores else 0,
        "std_dev": round(statistics.stdev(scores), 1) if len(scores) > 1 else 0,
        "grade_distribution": grade_dist
    }


def list_assignments_tool():
    """List all graded assignments with counts and averages."""
    rows = _load_master_csv()

    assignments = defaultdict(list)
    for row in rows:
        assignments[row["assignment"]].append(row)

    result = []
    for name, items in sorted(assignments.items()):
        scores = [r["score"] for r in items]
        unique_students = set(r["student_name"] for r in items)
        result.append({
            "assignment": name,
            "student_count": len(unique_students),
            "submission_count": len(scores),
            "average_score": round(sum(scores) / len(scores), 1) if scores else 0
        })

    # Also check saved assignment configs
    saved_configs = []
    if os.path.exists(ASSIGNMENTS_DIR):
        for f in os.listdir(ASSIGNMENTS_DIR):
            if f.endswith('.json'):
                saved_configs.append(f.replace('.json', ''))

    return {
        "graded_assignments": result,
        "saved_configs": saved_configs,
        "total_graded": len(result)
    }


def create_focus_assignment(name, category=None, points=None, date=None, description=None):
    """Launch Focus automation to create an assignment."""
    # Check credentials
    if not os.path.exists(CREDS_FILE):
        return {"error": "VPortal credentials not configured. Go to Settings > Tools to set them up."}

    script_path = os.path.join(PROJECT_ROOT, "focus-automation.js")
    if not os.path.exists(script_path):
        return {"error": "focus-automation.js not found in project root."}

    cmd = ["node", script_path, "assignment", "--name", name]
    if category:
        cmd.extend(["--category", category])
    if points:
        cmd.extend(["--points", str(points)])
    if date:
        cmd.extend(["--date", date])
    if description:
        cmd.extend(["--description", description])

    try:
        process = subprocess.Popen(
            cmd,
            cwd=PROJECT_ROOT,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE
        )
        return {
            "status": "launched",
            "message": f"Browser automation started for '{name}'. Check your phone for 2FA approval, then review the form before saving.",
            "pid": process.pid
        }
    except FileNotFoundError:
        return {"error": "Node.js not found. Make sure Node.js is installed."}
    except Exception as e:
        return {"error": f"Failed to launch automation: {str(e)}"}


def export_grades_csv(assignment=None, period=None):
    """Export grades as Focus-compatible CSV."""
    results = _load_results()
    if not results:
        return {"error": "No grading results to export"}

    os.makedirs(EXPORTS_DIR, exist_ok=True)

    # Filter
    filtered = results
    if assignment:
        filtered = [r for r in filtered
                    if assignment.lower() in r.get('assignment', '').lower()]
    if period:
        filtered = [r for r in filtered
                    if r.get('period', '') == period or r.get('quarter', '') == period]

    if not filtered:
        return {"error": "No results match the specified filters"}

    # Group by period
    by_period = defaultdict(list)
    for r in filtered:
        p = r.get('period', r.get('quarter', 'All'))
        by_period[p].append(r)

    safe_name = assignment or 'grades'
    safe_name = ''.join(c if c.isalnum() or c in ' -_' else '' for c in safe_name).strip().replace(' ', '_')

    exported_files = []
    total_rows = 0
    for p, items in by_period.items():
        safe_period = p.replace(' ', '_').replace('/', '-')
        filename = f"{safe_name}_{safe_period}.csv"
        filepath = os.path.join(EXPORTS_DIR, filename)

        csv_lines = ['Student ID,Score']
        matched = 0
        for r in items:
            student_id = r.get('student_id', '')
            score = r.get('score', 0)
            if student_id:
                csv_lines.append(f"{student_id},{score}")
                matched += 1
            else:
                name = r.get('student_name', 'Unknown')
                csv_lines.append(f"# {name},{score}")

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(csv_lines))

        exported_files.append({
            "file": filename,
            "period": p,
            "rows": matched,
            "download_url": "/api/download-export/" + filename,
        })
        total_rows += matched

    download_urls = [{"url": f["download_url"], "filename": f["file"]} for f in exported_files]

    result = {
        "status": "exported",
        "export_dir": EXPORTS_DIR,
        "files": exported_files,
        "total_rows": total_rows,
        "download_urls": download_urls,
    }

    # Single-file convenience: set top-level download_url for the simple download button path
    if len(exported_files) == 1:
        result["download_url"] = exported_files[0]["download_url"]
        result["filename"] = exported_files[0]["file"]

    return result


def analyze_grade_causes(assignment_name, period=None, score_threshold=None):
    """Deep analysis of what caused grades on an assignment."""
    results = _load_results()

    # Filter by assignment (partial match)
    matched = [r for r in results if assignment_name.lower() in r.get('assignment', '').lower()]
    if not matched:
        return {"error": f"No results found matching '{assignment_name}'"}

    actual_name = matched[0].get('assignment', '')
    matched = [r for r in results if r.get('assignment', '') == actual_name]

    if period:
        matched = [r for r in matched if r.get('period', '') == period]

    if score_threshold is not None:
        matched = [r for r in matched if _safe_int_score(r.get('score')) < score_threshold]

    if not matched:
        return {"error": "No results match the specified filters"}

    total = len(matched)
    scores = [_safe_int_score(r.get('score')) for r in matched]

    # Category breakdown analysis
    cat_totals = defaultdict(list)
    for r in matched:
        bd = r.get('breakdown', {})
        if bd:
            for cat, val in bd.items():
                cat_totals[cat].append(val)

    category_analysis = {}
    for cat, vals in cat_totals.items():
        avg = round(sum(vals) / len(vals), 1) if vals else 0
        zeros = sum(1 for v in vals if v == 0)
        category_analysis[cat] = {
            "average": avg,
            "zeros": zeros,
            "zero_pct": round(zeros / len(vals) * 100, 1) if vals else 0,
            "min": min(vals) if vals else 0,
            "max": max(vals) if vals else 0,
        }

    # Unanswered/omitted questions analysis
    unanswered_counts = defaultdict(int)
    students_with_unanswered = 0
    total_unanswered = 0
    for r in matched:
        uq = r.get('unanswered_questions')
        if uq and isinstance(uq, list) and len(uq) > 0:
            students_with_unanswered += 1
            total_unanswered += len(uq)
            for q in uq:
                unanswered_counts[q] += 1

    # Sort by frequency
    most_skipped = sorted(unanswered_counts.items(), key=lambda x: x[1], reverse=True)[:10]

    # Score impact of unanswered questions
    scores_with_unanswered = [_safe_int_score(r.get('score')) for r in matched
                              if r.get('unanswered_questions') and len(r.get('unanswered_questions', [])) > 0]
    scores_without_unanswered = [_safe_int_score(r.get('score')) for r in matched
                                 if not r.get('unanswered_questions') or len(r.get('unanswered_questions', [])) == 0]

    unanswered_impact = {}
    if scores_with_unanswered and scores_without_unanswered:
        avg_with = round(sum(scores_with_unanswered) / len(scores_with_unanswered), 1)
        avg_without = round(sum(scores_without_unanswered) / len(scores_without_unanswered), 1)
        unanswered_impact = {
            "students_with_omissions": len(scores_with_unanswered),
            "students_without_omissions": len(scores_without_unanswered),
            "avg_score_with_omissions": avg_with,
            "avg_score_without_omissions": avg_without,
            "score_gap": round(avg_without - avg_with, 1),
            "omission_pct": round(len(scores_with_unanswered) / total * 100, 1),
        }

    # Identify weakest category
    weakest_cat = None
    if category_analysis:
        weakest_cat = min(category_analysis.items(), key=lambda x: x[1]["average"])
        weakest_cat = {"name": weakest_cat[0], **weakest_cat[1]}

    return {
        "assignment_name": actual_name,
        "students_analyzed": total,
        "score_filter": f"below {score_threshold}" if score_threshold else "all students",
        "score_summary": {
            "mean": round(sum(scores) / len(scores), 1) if scores else 0,
            "median": round(statistics.median(scores), 1) if scores else 0,
        },
        "category_breakdown": category_analysis,
        "weakest_category": weakest_cat,
        "unanswered_questions": {
            "students_with_omissions": students_with_unanswered,
            "total_omissions": total_unanswered,
            "avg_omissions_per_student": round(total_unanswered / students_with_unanswered, 1) if students_with_unanswered else 0,
            "most_skipped_questions": [{"question": q, "times_skipped": c,
                                        "pct_students": round(c / total * 100, 1)}
                                       for q, c in most_skipped],
        },
        "omission_score_impact": unanswered_impact,
    }


def get_feedback_patterns(assignment_name, period=None):
    """Analyze feedback patterns across an assignment."""
    results = _load_results()

    matched = [r for r in results if assignment_name.lower() in r.get('assignment', '').lower()]
    if not matched:
        return {"error": f"No results found matching '{assignment_name}'"}

    actual_name = matched[0].get('assignment', '')
    matched = [r for r in results if r.get('assignment', '') == actual_name]

    if period:
        matched = [r for r in matched if r.get('period', '') == period]

    if not matched:
        return {"error": "No results match the specified filters"}

    # Aggregate strengths and developing skills
    all_strengths = []
    all_developing = []
    for r in matched:
        skills = r.get('skills_demonstrated', {})
        if isinstance(skills, dict):
            s = skills.get('strengths', [])
            d = skills.get('developing', [])
            if isinstance(s, list):
                all_strengths.extend(s)
            if isinstance(d, list):
                all_developing.extend(d)

    # Count frequency of skill mentions
    strength_freq = defaultdict(int)
    for s in all_strengths:
        strength_freq[s.strip()] += 1

    developing_freq = defaultdict(int)
    for d in all_developing:
        developing_freq[d.strip()] += 1

    top_strengths = sorted(strength_freq.items(), key=lambda x: x[1], reverse=True)[:8]
    top_developing = sorted(developing_freq.items(), key=lambda x: x[1], reverse=True)[:8]

    # Feedback content sampling (common phrases from low and high scorers)
    low_scorers = sorted(matched, key=lambda r: _safe_int_score(r.get('score')))[:5]
    high_scorers = sorted(matched, key=lambda r: _safe_int_score(r.get('score')), reverse=True)[:5]

    low_feedback_samples = []
    for r in low_scorers:
        fb = r.get('feedback', '')
        if fb:
            low_feedback_samples.append({
                "student": r.get('student_name', '').split()[0] if r.get('student_name') else 'Unknown',
                "score": _safe_int_score(r.get('score')),
                "feedback_preview": fb[:250] + "..." if len(fb) > 250 else fb,
                "unanswered_count": len(r.get('unanswered_questions', []) or []),
            })

    high_feedback_samples = []
    for r in high_scorers:
        fb = r.get('feedback', '')
        if fb:
            high_feedback_samples.append({
                "student": r.get('student_name', '').split()[0] if r.get('student_name') else 'Unknown',
                "score": _safe_int_score(r.get('score')),
                "feedback_preview": fb[:250] + "..." if len(fb) > 250 else fb,
            })

    # Marker status distribution
    marker_dist = defaultdict(int)
    for r in matched:
        marker_dist[r.get('marker_status', 'unknown')] += 1

    return {
        "assignment_name": actual_name,
        "total_students": len(matched),
        "common_strengths": [{"skill": s, "count": c, "pct": round(c / len(matched) * 100, 1)}
                             for s, c in top_strengths],
        "common_areas_for_growth": [{"skill": d, "count": c, "pct": round(c / len(matched) * 100, 1)}
                                     for d, c in top_developing],
        "lowest_scoring_feedback": low_feedback_samples,
        "highest_scoring_feedback": high_feedback_samples,
        "marker_status": dict(marker_dist),
    }


def compare_periods(assignment_name=None):
    """Compare performance across class periods."""
    results = _load_results()

    if assignment_name:
        results = [r for r in results if assignment_name.lower() in r.get('assignment', '').lower()]
        if results:
            actual_name = results[0].get('assignment', '')
            results = [r for r in results if r.get('assignment', '') == actual_name]
        else:
            return {"error": f"No results found matching '{assignment_name}'"}

    if not results:
        return {"error": "No grading results available"}

    # Load roster for authoritative student counts per period
    roster = _load_roster()
    roster_by_period = defaultdict(list)
    for s in roster:
        roster_by_period[s["period"]].append(s)

    # Group by period
    by_period = defaultdict(list)
    for r in results:
        p = r.get('period', 'Unknown')
        by_period[p].append(r)

    period_data = []
    for p, items in sorted(by_period.items()):
        scores = [_safe_int_score(r.get('score')) for r in items]

        # Count unique graded students and compute grade dist from per-student averages
        period_students = defaultdict(list)
        for r in items:
            sname = r.get('student_name', r.get('name', 'Unknown'))
            period_students[sname].append(_safe_int_score(r.get('score')))

        student_avg_scores = []
        for sname, s_scores in period_students.items():
            student_avg_scores.append(round(sum(s_scores) / len(s_scores), 1))

        grade_dist = {
            "A": len([s for s in student_avg_scores if s >= 90]),
            "B": len([s for s in student_avg_scores if 80 <= s < 90]),
            "C": len([s for s in student_avg_scores if 70 <= s < 80]),
            "D": len([s for s in student_avg_scores if 60 <= s < 70]),
            "F": len([s for s in student_avg_scores if s < 60])
        }

        # Category averages
        cat_totals = defaultdict(list)
        for r in items:
            bd = r.get('breakdown', {})
            if bd:
                for cat, val in bd.items():
                    cat_totals[cat].append(val)

        cat_avgs = {}
        for cat, vals in cat_totals.items():
            cat_avgs[cat] = round(sum(vals) / len(vals), 1) if vals else 0

        # Unanswered rate
        with_omissions = sum(1 for r in items
                             if r.get('unanswered_questions') and len(r.get('unanswered_questions', [])) > 0)

        # Use roster count (authoritative) if available, else graded count
        roster_match = roster_by_period.get(p, [])
        total_students = len(roster_match) if roster_match else len(period_students)

        period_data.append({
            "period": p,
            "student_count": total_students,
            "students_with_grades": len(period_students),
            "submission_count": len(items),
            "average": round(sum(scores) / len(scores), 1) if scores else 0,
            "median": round(statistics.median(scores), 1) if scores else 0,
            "grade_distribution": grade_dist,
            "category_averages": cat_avgs,
            "omission_rate": round(with_omissions / len(items) * 100, 1) if items else 0,
        })

    # Rank periods
    period_data.sort(key=lambda x: x["average"], reverse=True)

    return {
        "assignment": assignment_name or "all assignments",
        "periods": period_data,
        "best_period": period_data[0]["period"] if period_data else None,
        "lowest_period": period_data[-1]["period"] if period_data else None,
    }


def _load_period_class_levels():
    """Load class level (advanced/standard/support) for each period from metadata."""
    levels = {}
    if not os.path.exists(PERIODS_DIR):
        return levels
    for f in os.listdir(PERIODS_DIR):
        if f.endswith('.meta.json'):
            try:
                with open(os.path.join(PERIODS_DIR, f), 'r') as fh:
                    meta = json.load(fh)
                period_name = meta.get('period_name', f.replace('.csv.meta.json', ''))
                class_level = meta.get('class_level', 'standard')
                levels[period_name] = class_level
            except Exception:
                pass
    return levels


def _load_accommodations():
    """Load student accommodation data (IEP/504 presets)."""
    accommodations = {}
    if not os.path.exists(ACCOMMODATIONS_DIR):
        return accommodations
    for f in os.listdir(ACCOMMODATIONS_DIR):
        if f.endswith('.json'):
            try:
                with open(os.path.join(ACCOMMODATIONS_DIR, f), 'r') as fh:
                    data = json.load(fh)
                student_id = f.replace('.json', '')
                accommodations[student_id] = {
                    "presets": data.get('presets', []),
                    "notes": data.get('notes', ''),
                    "student_id": student_id,
                }
            except Exception:
                pass
    return accommodations


def _load_settings():
    """Load teacher settings (subject, state, grade level, AI notes)."""
    settings_path = os.path.expanduser("~/.graider_settings.json")
    if os.path.exists(settings_path):
        try:
            with open(settings_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _load_standards():
    """Load curriculum standards based on teacher's configured subject/state."""
    settings = _load_settings()
    config = settings.get('config', {})
    state = config.get('state', 'FL').lower()
    subject = config.get('subject', '').lower().replace(' ', '_')

    # Map subject names to filenames
    subject_map = {
        'us_history': 'us_history',
        'u.s._history': 'us_history',
        'world_history': 'world_history',
        'civics': 'civics',
        'geography': 'geography',
        'english/ela': 'english-ela',
        'english': 'english-ela',
        'ela': 'english-ela',
        'math': 'math',
        'science': 'science',
        'social_studies': 'social_studies',
    }
    subject_key = subject_map.get(subject, subject)
    filename = f"standards_fl_{subject_key}.json"
    filepath = os.path.join(STANDARDS_DIR, filename)

    if os.path.exists(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return data.get('standards', [])
        except Exception:
            pass
    return []


def _load_saved_lessons():
    """Load saved lesson plan titles and topics."""
    lessons = []
    if not os.path.exists(LESSONS_DIR):
        return lessons
    for unit_dir in os.listdir(LESSONS_DIR):
        unit_path = os.path.join(LESSONS_DIR, unit_dir)
        if os.path.isdir(unit_path):
            for f in os.listdir(unit_path):
                if f.endswith('.json'):
                    try:
                        with open(os.path.join(unit_path, f), 'r') as fh:
                            data = json.load(fh)
                        lessons.append({
                            "title": data.get('title', f.replace('.json', '')),
                            "unit": unit_dir,
                            "standards": data.get('standards', []),
                        })
                    except Exception:
                        pass
    return lessons


def _analyze_group_weaknesses(group_results):
    """Shared weakness analysis for a group of student results.
    Returns dict with category_weaknesses, content_gaps, developing_skills, strengths, scores."""
    total = len(group_results)
    if not total:
        return None

    # Category breakdown
    cat_totals = defaultdict(list)
    for r in group_results:
        bd = r.get('breakdown', {})
        if bd:
            for cat, val in bd.items():
                try:
                    cat_totals[cat].append(int(float(val)) if val else 0)
                except (ValueError, TypeError):
                    pass

    category_weaknesses = []
    for cat, vals in cat_totals.items():
        avg = round(sum(vals) / len(vals), 1) if vals else 0
        max_possible = max(vals) if vals else 0
        zeros = sum(1 for v in vals if v == 0)
        category_weaknesses.append({
            "category": cat,
            "average": avg,
            "max_seen": max_possible,
            "zero_count": zeros,
            "zero_pct": round(zeros / len(vals) * 100, 1) if vals else 0,
        })
    category_weaknesses.sort(key=lambda x: x["average"])

    # Unanswered questions
    unanswered_counts = defaultdict(int)
    for r in group_results:
        uq = r.get('unanswered_questions')
        if uq and isinstance(uq, list):
            for q in uq:
                unanswered_counts[q] += 1
    content_gaps = sorted(unanswered_counts.items(), key=lambda x: x[1], reverse=True)[:8]

    # Skills
    developing_freq = defaultdict(int)
    strength_freq = defaultdict(int)
    for r in group_results:
        skills = r.get('skills_demonstrated', {})
        if isinstance(skills, dict):
            for s in (skills.get('developing', []) or []):
                developing_freq[s.strip()] += 1
            for s in (skills.get('strengths', []) or []):
                strength_freq[s.strip()] += 1

    top_developing = sorted(developing_freq.items(), key=lambda x: x[1], reverse=True)[:6]
    top_strengths = sorted(strength_freq.items(), key=lambda x: x[1], reverse=True)[:4]

    # Scores
    scores = [_safe_int_score(r.get('score')) for r in group_results]
    failing = [s for s in scores if s < 70]
    avg_score = round(sum(scores) / len(scores), 1) if scores else 0

    # Omissions
    omission_count = sum(1 for r in group_results
                         if r.get('unanswered_questions') and len(r.get('unanswered_questions', [])) > 0)

    return {
        "total_students": total,
        "average_score": avg_score,
        "failing_count": len(failing),
        "failing_pct": round(len(failing) / total * 100, 1) if total else 0,
        "omission_rate": round(omission_count / total * 100, 1) if total else 0,
        "category_weaknesses": category_weaknesses,
        "content_gaps": [{"topic": q, "students_missed": c,
                          "pct": round(c / total * 100, 1)}
                         for q, c in content_gaps],
        "developing_skills": [{"skill": s, "count": c} for s, c in top_developing],
        "student_strengths": [{"skill": s, "count": c} for s, c in top_strengths],
    }


def _match_standards(weakness_data, standards, target_dok=None):
    """Match curriculum standards to identified weaknesses, optionally filtering by DOK level."""
    if not standards or not weakness_data:
        return []

    weakness_keywords = set()
    for gap in weakness_data.get("content_gaps", []):
        for word in gap["topic"].lower().split():
            if len(word) > 3:
                weakness_keywords.add(word)
    for dev in weakness_data.get("developing_skills", []):
        for word in dev["skill"].lower().split():
            if len(word) > 3:
                weakness_keywords.add(word)

    relevant = []
    for std in standards:
        topics = [t.lower() for t in std.get('topics', [])]
        benchmark = std.get('benchmark', '').lower()
        vocab = [v.lower() for v in std.get('vocabulary', [])]
        all_text = ' '.join(topics + [benchmark] + vocab)

        match_count = sum(1 for kw in weakness_keywords if kw in all_text)
        if match_count > 0:
            std_dok = std.get('dok', '')
            dok_match = True
            if target_dok is not None and std_dok:
                try:
                    dok_val = int(std_dok) if str(std_dok).isdigit() else 0
                    dok_match = dok_val <= target_dok
                except (ValueError, TypeError):
                    dok_match = True

            relevant.append({
                "code": std.get('code', ''),
                "benchmark": std.get('benchmark', '')[:200],
                "topics": std.get('topics', []),
                "dok": std_dok,
                "essential_questions": std.get('essential_questions', [])[:2],
                "learning_targets": std.get('learning_targets', [])[:2],
                "relevance_score": match_count,
                "dok_appropriate": dok_match,
            })

    relevant.sort(key=lambda x: (-int(x["dok_appropriate"]), -x["relevance_score"]))
    return relevant[:5]


def recommend_next_lesson(assignment_name=None, period=None, num_assignments=1):
    """Analyze performance and recommend next lesson focus with period differentiation and IEP awareness."""
    results = _load_results()
    if not results:
        return {"error": "No grading results available"}

    num_assignments = min(num_assignments or 1, 5)

    # Determine which assignment(s) to analyze
    if assignment_name:
        target_results = [r for r in results
                          if assignment_name.lower() in r.get('assignment', '').lower()]
    else:
        sorted_results = sorted(results, key=lambda r: r.get('graded_at', ''), reverse=True)
        recent_assignments = []
        seen = set()
        for r in sorted_results:
            a = r.get('assignment', '')
            if a and a not in seen:
                seen.add(a)
                recent_assignments.append(a)
            if len(recent_assignments) >= num_assignments:
                break
        target_results = [r for r in results if r.get('assignment', '') in recent_assignments]

    if period:
        target_results = [r for r in target_results if r.get('period', '') == period]

    if not target_results:
        return {"error": "No matching results found for analysis"}

    analyzed_assignments = list(set(r.get('assignment', '') for r in target_results))

    # Load supplementary data
    period_levels = _load_period_class_levels()
    accommodations = _load_accommodations()
    standards = _load_standards()
    settings = _load_settings()
    config = settings.get('config', {})
    global_notes = settings.get('globalAINotes', '')
    saved_lessons = _load_saved_lessons()

    # === Overall weakness analysis ===
    overall = _analyze_group_weaknesses(target_results)

    # === Per-class-level analysis ===
    # Group results by class level (advanced/standard/support)
    level_groups = defaultdict(list)
    periods_by_level = defaultdict(list)

    for r in target_results:
        p = r.get('period', '')
        level = period_levels.get(p, 'standard')
        level_groups[level].append(r)
        if p not in periods_by_level[level]:
            periods_by_level[level].append(p)

    # DOK targets by class level
    dok_targets = {"advanced": 4, "standard": 3, "support": 2}

    class_level_analysis = {}
    for level in ["advanced", "standard", "support"]:
        group = level_groups.get(level, [])
        if not group:
            continue
        weakness = _analyze_group_weaknesses(group)
        if not weakness:
            continue

        target_dok = dok_targets.get(level, 3)
        matched_standards = _match_standards(weakness, standards, target_dok=target_dok)

        # Build skill needs for this level
        skill_needs = []
        cw = weakness["category_weaknesses"]
        if cw:
            weakest = cw[0]
            skill_needs.append(f"Weakest area: {weakest['category']} (avg {weakest['average']})")
        cg = weakness["content_gaps"]
        if cg:
            skill_needs.append(f"Most skipped: '{cg[0]['topic']}' ({cg[0]['pct']}%)")
        ds = weakness["developing_skills"]
        if ds:
            skill_needs.append(f"Top developing: '{ds[0]['skill']}' ({ds[0]['count']} students)")
        if weakness["omission_rate"] > 40:
            skill_needs.append(f"High omission rate: {weakness['omission_rate']}%")

        class_level_analysis[level] = {
            "class_level": level,
            "periods": periods_by_level[level],
            "target_dok": target_dok,
            "identified_needs": skill_needs,
            "recommended_standards": matched_standards,
            **weakness,
        }

    # === IEP/504 Accommodation Analysis ===
    accommodation_analysis = None
    if accommodations:
        # Match accommodated students to results by student_id or name
        accom_ids = set(accommodations.keys())
        accom_results = []
        non_accom_results = []
        matched_students = {}

        for r in target_results:
            sid = r.get('student_id', '')
            sname = r.get('student_name', '').lower().replace(' ', '_')
            if sid in accom_ids:
                accom_results.append(r)
                matched_students[sid] = accommodations[sid]
            elif sname in accom_ids:
                accom_results.append(r)
                matched_students[sname] = accommodations[sname]
            else:
                non_accom_results.append(r)

        if accom_results:
            accom_weakness = _analyze_group_weaknesses(accom_results)
            non_accom_weakness = _analyze_group_weaknesses(non_accom_results) if non_accom_results else None

            # Collect unique presets across matched students
            all_presets = set()
            for info in matched_students.values():
                for p in info.get('presets', []):
                    all_presets.add(p)

            score_gap = None
            if accom_weakness and non_accom_weakness:
                score_gap = round(non_accom_weakness["average_score"] - accom_weakness["average_score"], 1)

            accommodation_analysis = {
                "accommodated_student_count": len(accom_results),
                "total_student_count": len(target_results),
                "accommodation_presets_in_use": sorted(all_presets),
                "accommodated_avg_score": accom_weakness["average_score"] if accom_weakness else None,
                "non_accommodated_avg_score": non_accom_weakness["average_score"] if non_accom_weakness else None,
                "score_gap": score_gap,
                "accommodated_failing_pct": accom_weakness["failing_pct"] if accom_weakness else None,
                "accommodated_omission_rate": accom_weakness["omission_rate"] if accom_weakness else None,
                "accommodated_weaknesses": accom_weakness["category_weaknesses"][:3] if accom_weakness else [],
                "note": "IEP/504 students may need modified lesson pacing, scaffolded activities, or alternative assessments.",
            }

    # === Overall standards (no DOK filter) for the combined view ===
    overall_standards = _match_standards(overall, standards, target_dok=None)

    # === Build top-level identified needs ===
    skill_needs = []
    if overall["category_weaknesses"]:
        weakest = overall["category_weaknesses"][0]
        skill_needs.append(f"Weakest rubric area: {weakest['category']} (avg {weakest['average']})")
    if overall["content_gaps"]:
        top_gap = overall["content_gaps"][0]
        skill_needs.append(f"Most skipped section: '{top_gap['topic']}' ({top_gap['students_missed']} students, "
                          f"{top_gap['pct']}%)")
    if overall["developing_skills"]:
        skill_needs.append(f"Top developing skill: '{overall['developing_skills'][0]['skill']}' "
                          f"({overall['developing_skills'][0]['count']} students)")
    if overall["omission_rate"] > 40:
        skill_needs.append(f"High omission rate: {overall['omission_rate']}% of students "
                          f"left questions blank — consider assignment completion mini-lesson")

    # Note if class levels diverge significantly
    if len(class_level_analysis) > 1:
        level_avgs = {lv: d["average_score"] for lv, d in class_level_analysis.items()}
        max_avg = max(level_avgs.values())
        min_avg = min(level_avgs.values())
        if max_avg - min_avg > 10:
            skill_needs.append(
                f"Large gap between class levels: {max_avg} (highest) vs {min_avg} (lowest) — "
                f"differentiated lesson planning recommended"
            )

    return {
        "analyzed_assignments": analyzed_assignments,
        "period_filter": period or "all periods",
        "total_students": overall["total_students"],
        "average_score": overall["average_score"],
        "failing_count": overall["failing_count"],
        "failing_pct": overall["failing_pct"],
        "identified_needs": skill_needs,
        "category_weaknesses": overall["category_weaknesses"],
        "content_gaps": overall["content_gaps"],
        "developing_skills": overall["developing_skills"],
        "student_strengths": overall["student_strengths"],
        "relevant_standards": overall_standards,
        "existing_lessons": [l["title"] for l in saved_lessons],
        "teacher_subject": config.get('subject', ''),
        "teacher_grade": config.get('grade_level', ''),
        "period_differentiation_notes": global_notes[:500] if global_notes else "",
        "class_level_breakdown": class_level_analysis if class_level_analysis else None,
        "period_class_levels": period_levels if period_levels else None,
        "accommodation_analysis": accommodation_analysis,
    }


def _load_roster():
    """Load student roster from ~/.graider_data/periods/. Returns list of dicts with name, id, local_id, grade, period, course_codes."""
    roster = []
    if not os.path.exists(PERIODS_DIR):
        return roster
    # Load period metadata for course codes
    period_meta = {}
    for f in os.listdir(PERIODS_DIR):
        if f.endswith('.meta.json'):
            try:
                with open(os.path.join(PERIODS_DIR, f), 'r') as fh:
                    meta = json.load(fh)
                csv_name = f.replace('.meta.json', '')
                period_meta[csv_name] = meta
            except Exception:
                pass

    for f in sorted(os.listdir(PERIODS_DIR)):
        if not f.endswith('.csv'):
            continue
        meta = period_meta.get(f, {})
        period_name = meta.get('period_name', f.replace('.csv', '').replace('_', ' '))
        course_codes = meta.get('course_codes', [])
        filepath = os.path.join(PERIODS_DIR, f)
        try:
            with open(filepath, 'r', encoding='utf-8') as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    raw_name = row.get('Student', '').strip().strip('"')
                    student_id = row.get('Student ID', '').strip().strip('"')
                    local_id = row.get('Local ID', '').strip().strip('"')
                    grade = row.get('Grade', '').strip().strip('"')
                    # Convert "Last, First Middle" or "Last; First Middle" to "First Middle Last"
                    if ';' in raw_name:
                        parts = raw_name.split(';', 1)
                        last = parts[0].strip()
                        first = parts[1].strip() if len(parts) > 1 else ''
                        display_name = f"{first} {last}".strip()
                    elif ',' in raw_name:
                        parts = raw_name.split(',', 1)
                        last = parts[0].strip()
                        first = parts[1].strip() if len(parts) > 1 else ''
                        display_name = f"{first} {last}".strip()
                    else:
                        display_name = raw_name
                    if display_name:
                        roster.append({
                            "name": display_name,
                            "student_id": student_id,
                            "local_id": local_id,
                            "grade": grade,
                            "period": period_name,
                            "course_codes": course_codes,
                        })
        except Exception:
            pass
    return roster


def _load_parent_contacts():
    """Load parent contacts keyed by student ID."""
    if not os.path.exists(PARENT_CONTACTS_FILE):
        return {}
    try:
        with open(PARENT_CONTACTS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def lookup_student_info(student_name=None, student_id=None, student_ids=None, period=None):
    """Look up student roster and contact information.
    Supports batch lookup via student_ids (list of IDs)."""
    roster = _load_roster()
    parent_contacts = _load_parent_contacts()
    results_json = _load_results()

    # Build email lookup from grading results (student_id -> email)
    email_lookup = {}
    for r in results_json:
        sid = r.get('student_id', '')
        email = r.get('student_email', '')
        if sid and email:
            email_lookup[sid] = email

    if not roster and not parent_contacts:
        return {"error": "No student roster data found. Import from Focus SIS or upload class period CSVs in Settings."}

    # Batch lookup by student_ids list
    if student_ids and isinstance(student_ids, list):
        id_set = set(str(sid) for sid in student_ids)
        matches = [s for s in roster if s["student_id"] in id_set]
        # Also check parent contacts for IDs not found in roster
        found_ids = set(s["student_id"] for s in matches)
        for sid in id_set - found_ids:
            contact = parent_contacts.get(sid)
            if contact:
                matches.append({
                    "name": contact.get("student_name", "Unknown"),
                    "student_id": sid,
                    "local_id": "",
                    "grade": "",
                    "period": contact.get("period", ""),
                })
    else:
        # Single lookup mode
        matches = roster
        if student_id:
            matches = [s for s in matches if s["student_id"] == str(student_id)]
        if student_name:
            matches = [s for s in matches if _fuzzy_name_match(student_name, s["name"])]
        if period:
            # Normalize period input: "1" -> matches "Period 1", "Period 1" -> matches "Period 1"
            period_lower = period.lower().strip()
            matches = [s for s in matches
                       if period_lower in s["period"].lower()
                       or (period_lower.isdigit() and f"period {period_lower}" in s["period"].lower())]

        if not matches and student_id:
            # Try parent contacts as fallback (has student_id keys even without roster)
            contact = parent_contacts.get(str(student_id))
            if contact:
                matches = [{
                    "name": contact.get("student_name", "Unknown"),
                    "student_id": str(student_id),
                    "local_id": "",
                    "grade": "",
                    "period": contact.get("period", ""),
                }]

        if not matches and student_name:
            # Try parent contacts by name (fuzzy word match)
            for sid, contact in parent_contacts.items():
                if _fuzzy_name_match(student_name, contact.get("student_name", "")):
                    matches.append({
                        "name": contact.get("student_name", "Unknown"),
                        "student_id": sid,
                        "local_id": "",
                        "grade": "",
                        "period": contact.get("period", ""),
                    })

    if not matches:
        return {"error": "No students found matching the search criteria.", "searched": {
            "name": student_name, "id": student_id, "ids": student_ids, "period": period
        }}

    # Deduplicate by student_id (batch mode may have overlap)
    seen_ids = set()
    unique_matches = []
    for s in matches:
        sid = s["student_id"]
        if sid not in seen_ids:
            seen_ids.add(sid)
            unique_matches.append(s)

    # Enrich each match with contact info, email, 504 status, schedule, and course codes
    students = []
    for s in unique_matches:
        sid = s["student_id"]
        contact = parent_contacts.get(sid, {})
        student_email = email_lookup.get(sid, "")

        entry = {
            "name": s["name"],
            "student_id": sid,
            "local_id": s.get("local_id", ""),
            "grade_level": s.get("grade", ""),
            "period": s["period"],
            "course_codes": s.get("course_codes", []),
            "student_email": student_email,
            "parent_emails": contact.get("parent_emails", []),
            "parent_phones": contact.get("parent_phones", []),
            "has_504": contact.get("has_504", False),
            "contacts": contact.get("contacts", []),
            "schedule": contact.get("schedule", []),
        }
        students.append(entry)

    return {
        "students": students,
        "total_found": len(students),
    }


def generate_worksheet_tool(title, worksheet_type, vocab_terms=None, questions=None,
                            summary_prompt=None, summary_key_points=None,
                            total_points=100, style_name=None):
    """Generate a .docx worksheet and save to Grading Setup."""
    try:
        from backend.services.worksheet_generator import generate_worksheet
        # Load subject from teacher settings
        settings = _load_settings()
        config = settings.get('config', {})
        subject = config.get('subject', '')
        return generate_worksheet(
            title=title,
            worksheet_type=worksheet_type,
            vocab_terms=vocab_terms,
            questions=questions,
            summary_prompt=summary_prompt,
            summary_key_points=summary_key_points,
            total_points=total_points,
            subject=subject,
            style_name=style_name
        )
    except ImportError:
        return {"error": "python-docx not installed. Run: pip install python-docx"}
    except Exception as e:
        return {"error": "Failed to generate worksheet: " + str(e)}


def generate_document_tool(title, content, style_name=None, save_to_builder=False):
    """Generate a formatted Word document with rich typography."""
    try:
        from backend.services.document_generator import generate_document
        return generate_document(
            title=title, content=content,
            style_name=style_name, save_to_builder=save_to_builder
        )
    except ImportError:
        return {"error": "python-docx not installed. Run: pip install python-docx"}
    except Exception as e:
        return {"error": "Failed to generate document: " + str(e)}


def generate_csv_tool(filename, headers, rows):
    """Generate a downloadable CSV or XLSX file based on filename extension."""
    from urllib.parse import quote

    EXPORT_DIR = os.path.expanduser("~/Downloads/Graider/Exports")
    os.makedirs(EXPORT_DIR, exist_ok=True)

    # Sanitize filename
    safe_name = "".join(c for c in filename if c.isalnum() or c in ' -_.').strip()
    is_xlsx = safe_name.lower().endswith('.xlsx')

    # Default to .xlsx if no recognized extension
    if not safe_name.lower().endswith(('.csv', '.xlsx')):
        safe_name += '.xlsx'
        is_xlsx = True

    filepath = os.path.join(EXPORT_DIR, safe_name)

    if is_xlsx:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quiz"

        # Write header row with styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Auto-fit column widths (approximate)
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in rows:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

        wb.save(filepath)
    else:
        import csv
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)

    download_url = "/api/download-csv/" + quote(safe_name)

    return {
        "status": "created",
        "filename": safe_name,
        "filepath": filepath,
        "download_url": download_url,
        "row_count": len(rows),
        "columns": headers,
    }


def save_document_style_tool(name, style):
    """Save a named visual style for documents."""
    try:
        from backend.services.document_generator import save_style
        return save_style(name=name, style_dict=style)
    except Exception as e:
        return {"error": "Failed to save style: " + str(e)}


def list_document_styles_tool():
    """List saved document visual styles."""
    try:
        from backend.services.document_generator import list_styles
        return list_styles()
    except Exception as e:
        return {"error": "Failed to list styles: " + str(e)}


def get_missing_assignments(student_name=None, period=None, assignment_name=None):
    """Find missing/unsubmitted assignments.

    Three modes:
    1. student_name → list assignments this student hasn't submitted
    2. period → list all students in that period with missing work
    3. assignment_name → list students who are missing that specific assignment
    """
    rows = _load_master_csv()
    if not rows:
        return {"error": "No grading data available"}

    period_assigns, assign_display, merge_map = _get_period_assignments(rows)

    # Build per-student data keyed by student_id (merges name variants like "vincent scar" + "Vincent Ray Scarola")
    student_data = defaultdict(lambda: {"assigns": set(), "period": "", "name": ""})
    for r in rows:
        sid = r.get("student_id", "")
        name = r["student_name"]
        if not sid or sid == "UNKNOWN":
            continue
        p = _normalize_period(r.get("period", "") or r.get("quarter", ""))
        norm_assign = _normalize_assignment_name(r["assignment"])
        norm_assign = merge_map.get(norm_assign, norm_assign)
        student_data[sid]["assigns"].add(norm_assign)
        if p:
            student_data[sid]["period"] = p
        # Keep the longest name variant as display name
        if len(name) > len(student_data[sid]["name"]):
            student_data[sid]["name"] = name

    # Mode 1: Specific student
    if student_name:
        matches = [(sid, d) for sid, d in student_data.items() if student_name.lower() in d["name"].lower()]
        if not matches:
            return {"error": f"No student found matching '{student_name}'"}
        sid, data = matches[0]
        p = data["period"]
        all_assigns = period_assigns.get(p, set())
        missing = all_assigns - data["assigns"]
        return {
            "student_name": data["name"],
            "period": p,
            "submitted_count": len(data["assigns"]),
            "total_in_period": len(all_assigns),
            "missing_count": len(missing),
            "submitted": sorted(assign_display.get(n, n) for n in data["assigns"]),
            "missing": sorted(assign_display.get(n, n) for n in missing),
        }

    # Mode 2: By period — all students missing work
    if period:
        period_norm = _normalize_period(period)
        all_assigns = period_assigns.get(period_norm, set())
        if not all_assigns:
            return {"error": f"No graded assignments found for {period_norm}"}

        students_missing = []
        for sid, data in student_data.items():
            if data["period"] != period_norm:
                continue
            missing = all_assigns - data["assigns"]
            if missing:
                students_missing.append({
                    "student_name": data["name"],
                    "missing_count": len(missing),
                    "submitted_count": len(data["assigns"]),
                    "missing": sorted(assign_display.get(n, n) for n in missing),
                })
        students_missing.sort(key=lambda x: -x["missing_count"])

        return {
            "period": period_norm,
            "total_assignments": len(all_assigns),
            "students_with_missing": len(students_missing),
            "students": students_missing[:30],
        }

    # Mode 3: By assignment — which students are missing it
    if assignment_name:
        target_norm = _normalize_assignment_name(assignment_name)
        # Find matching assignment across all periods
        found_in_periods = set()
        for p, assigns in period_assigns.items():
            for a in assigns:
                if assignment_name.lower() in assign_display.get(a, a).lower() or assignment_name.lower() in a.lower():
                    target_norm = a
                    found_in_periods.add(p)

        if not found_in_periods:
            return {"error": f"No assignment found matching '{assignment_name}'"}

        display_name = assign_display.get(target_norm, assignment_name)
        missing_students = []
        submitted_students = []
        for sid, data in student_data.items():
            if data["period"] not in found_in_periods:
                continue
            if target_norm in data["assigns"]:
                submitted_students.append({"student_name": data["name"], "period": data["period"]})
            else:
                missing_students.append({"student_name": data["name"], "period": data["period"]})

        missing_students.sort(key=lambda x: (x["period"], x["student_name"]))
        return {
            "assignment": display_name,
            "periods_assigned": sorted(found_in_periods),
            "submitted_count": len(submitted_students),
            "missing_count": len(missing_students),
            "missing_students": missing_students,
        }

    return {"error": "Provide student_name, period, or assignment_name to search"}


# ═══════════════════════════════════════════════════════
# STANDARDS LOOKUP
# ═══════════════════════════════════════════════════════

def get_standards_tool(topic=None, dok_max=None):
    """Look up curriculum standards filtered by topic and DOK level."""
    all_standards = _load_standards()
    if not all_standards:
        settings = _load_settings()
        config = settings.get('config', {})
        subj = config.get('subject', 'unknown')
        st = config.get('state', 'unknown')
        return {"error": f"No standards found for {subj} in {st}. Check Settings > Subject and State."}

    results = all_standards

    # Filter by DOK level
    if dok_max is not None:
        results = [s for s in results if s.get('dok', 99) <= dok_max]

    # Filter by topic keyword
    if topic:
        topic_lower = topic.lower()
        filtered = []
        for s in results:
            searchable = " ".join([
                s.get('benchmark', ''),
                " ".join(s.get('topics', [])),
                " ".join(s.get('vocabulary', [])),
                s.get('item_specs', ''),
            ]).lower()
            if topic_lower in searchable:
                filtered.append(s)
        results = filtered

    if not results:
        return {"error": f"No standards found matching '{topic or 'all'}' (DOK <= {dok_max or 'any'})"}

    return {
        "count": len(results),
        "standards": [
            {
                "code": s.get("code", ""),
                "benchmark": s.get("benchmark", ""),
                "dok": s.get("dok"),
                "topics": s.get("topics", []),
                "vocabulary": s.get("vocabulary", []),
                "learning_targets": s.get("learning_targets", []),
                "essential_questions": s.get("essential_questions", []),
                "sample_assessment": s.get("sample_assessment", ""),
            }
            for s in results
        ]
    }


# ═══════════════════════════════════════════════════════
# RECENT LESSONS
# ═══════════════════════════════════════════════════════

def get_recent_lessons(unit_name=None):
    """List saved lesson plans with full detail for document generation context."""
    if not os.path.exists(LESSONS_DIR):
        return {"error": "No saved lessons found. Generate and save lesson plans in the Planner tab first."}

    lessons = []
    for unit_dir in os.listdir(LESSONS_DIR):
        unit_path = os.path.join(LESSONS_DIR, unit_dir)
        if not os.path.isdir(unit_path):
            continue

        # Filter by unit name if provided
        if unit_name and unit_name.lower() not in unit_dir.lower():
            continue

        for fname in os.listdir(unit_path):
            if not fname.endswith('.json'):
                continue
            try:
                with open(os.path.join(unit_path, fname), 'r', encoding='utf-8') as fh:
                    data = json.load(fh)

                # Extract day-level details
                days_summary = []
                all_vocab = []
                all_standards = []
                for day in data.get('days', []):
                    day_info = {
                        "day": day.get("day"),
                        "topic": day.get("topic", ""),
                        "objective": day.get("objective", ""),
                    }
                    # Collect standards addressed
                    stds = day.get("standards_addressed", [])
                    if stds:
                        day_info["standards"] = stds
                        all_standards.extend(stds)
                    # Collect vocabulary
                    vocab = day.get("vocabulary", [])
                    for v in vocab:
                        term = v.get("term", v) if isinstance(v, dict) else str(v)
                        if term and term not in all_vocab:
                            all_vocab.append(term)
                    days_summary.append(day_info)

                lessons.append({
                    "unit": unit_dir,
                    "title": data.get("title", fname.replace(".json", "")),
                    "overview": data.get("overview", ""),
                    "essential_questions": data.get("essential_questions", []),
                    "num_days": len(data.get("days", [])),
                    "days": days_summary,
                    "vocabulary": all_vocab,
                    "standards_covered": list(set(all_standards)),
                    "saved_at": data.get("_saved_at", ""),
                })
            except Exception:
                continue

    if not lessons:
        msg = f"No lessons found for unit '{unit_name}'." if unit_name else "No saved lessons found."
        return {"error": msg}

    # Sort by saved_at (most recent first), cap at 10
    lessons.sort(key=lambda x: x.get("saved_at", ""), reverse=True)
    lessons = lessons[:10]

    # Group by unit for readability
    units = {}
    for lesson in lessons:
        unit = lesson["unit"]
        if unit not in units:
            units[unit] = []
        units[unit].append(lesson)

    return {
        "total_lessons": len(lessons),
        "units": list(units.keys()),
        "lessons": lessons,
    }


# ═══════════════════════════════════════════════════════
# CALENDAR TOOLS
# ═══════════════════════════════════════════════════════

CALENDAR_FILE = os.path.expanduser("~/.graider_data/teaching_calendar.json")


def _load_calendar():
    """Load calendar data from disk."""
    if not os.path.exists(CALENDAR_FILE):
        return {"scheduled_lessons": [], "holidays": [], "school_days": {}}
    try:
        with open(CALENDAR_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {"scheduled_lessons": [], "holidays": [], "school_days": {}}


def _save_calendar(data):
    """Persist calendar data to disk."""
    os.makedirs(os.path.dirname(CALENDAR_FILE), exist_ok=True)
    with open(CALENDAR_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)


def get_calendar(start_date=None, end_date=None):
    """Read the teaching calendar for a date range."""
    cal = _load_calendar()

    if not start_date:
        start_date = datetime.now().strftime('%Y-%m-%d')
    if not end_date:
        from datetime import timedelta
        end_dt = datetime.strptime(start_date, '%Y-%m-%d') + timedelta(days=7)
        end_date = end_dt.strftime('%Y-%m-%d')

    # Filter lessons in range
    lessons = [s for s in cal.get("scheduled_lessons", [])
               if start_date <= s.get("date", "") <= end_date]
    lessons.sort(key=lambda s: s.get("date", ""))

    # Filter holidays in range (including multi-day overlaps)
    holidays = []
    for h in cal.get("holidays", []):
        h_start = h.get("date", "")
        h_end = h.get("end_date", h_start)
        if h_end >= start_date and h_start <= end_date:
            holidays.append(h)

    return {
        "start_date": start_date,
        "end_date": end_date,
        "scheduled_lessons": lessons,
        "holidays": holidays,
        "total_lessons": len(lessons),
        "total_holidays": len(holidays),
    }


def schedule_lesson_tool(date, lesson_title, unit=None, day_number=None, lesson_file=None):
    """Schedule a lesson on the teaching calendar."""
    import uuid as _uuid
    if not date or not lesson_title:
        return {"error": "date and lesson_title are required"}

    cal = _load_calendar()

    # Pick a color based on unit name
    unit_colors = ['#6366f1', '#8b5cf6', '#ec4899', '#f59e0b', '#22c55e', '#06b6d4', '#ef4444']
    color_idx = hash(unit or '') % len(unit_colors)

    entry = {
        "id": str(_uuid.uuid4()),
        "date": date,
        "unit": unit or "",
        "lesson_title": lesson_title,
        "day_number": day_number,
        "lesson_file": lesson_file or "",
        "color": unit_colors[color_idx],
    }

    cal["scheduled_lessons"].append(entry)
    _save_calendar(cal)

    return {"status": "scheduled", "entry": entry}


def add_calendar_holiday(date, name, end_date=None):
    """Add a holiday or break to the teaching calendar."""
    if not date or not name:
        return {"error": "date and name are required"}

    cal = _load_calendar()

    holiday = {"date": date, "name": name}
    if end_date:
        holiday["end_date"] = end_date

    # Remove existing holiday on same date to avoid duplicates
    cal["holidays"] = [h for h in cal["holidays"] if h["date"] != date]
    cal["holidays"].append(holiday)
    cal["holidays"].sort(key=lambda h: h["date"])

    _save_calendar(cal)
    return {"status": "added", "holiday": holiday}


# ═══════════════════════════════════════════════════════
# PERSISTENT MEMORY
# ═══════════════════════════════════════════════════════

def _load_memories():
    """Load saved memories from disk. Returns list of fact strings."""
    if not os.path.exists(MEMORY_FILE):
        return []
    try:
        with open(MEMORY_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _save_memories(memories):
    """Persist memories list to disk."""
    os.makedirs(os.path.dirname(MEMORY_FILE), exist_ok=True)
    with open(MEMORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(memories, f, indent=2)


def save_memory(fact):
    """Save a fact to persistent memory for future conversations."""
    if not fact or not fact.strip():
        return {"error": "No fact provided"}

    fact = fact.strip()
    memories = _load_memories()

    # Duplicate check — skip if a very similar fact already exists
    fact_lower = fact.lower()
    for existing in memories:
        existing_text = existing.get("fact", existing) if isinstance(existing, dict) else str(existing)
        if existing_text.lower() == fact_lower:
            return {"status": "already_saved", "fact": fact, "message": "This fact is already saved."}

    # Build entry with timestamp
    entry = {"fact": fact, "saved_at": datetime.now().isoformat()}
    memories.append(entry)

    # Cap at MAX_MEMORIES (remove oldest if exceeded)
    if len(memories) > MAX_MEMORIES:
        memories = memories[-MAX_MEMORIES:]

    _save_memories(memories)
    return {"status": "saved", "fact": fact, "total_memories": len(memories)}


# ═══════════════════════════════════════════════════════
# STANDARDS INDEX
# ═══════════════════════════════════════════════════════

def list_all_standards_tool():
    """Return a compact index of ALL curriculum standards for the teacher's subject."""
    all_standards = _load_standards()
    if not all_standards:
        settings = _load_settings()
        config = settings.get('config', {})
        subj = config.get('subject', 'unknown')
        st = config.get('state', 'unknown')
        return {"error": f"No standards found for {subj} in {st}. Check Settings > Subject and State."}

    settings = _load_settings()
    config = settings.get('config', {})

    compact = []
    for s in all_standards:
        benchmark = s.get("benchmark", "")
        compact.append({
            "code": s.get("code", ""),
            "benchmark": benchmark[:120] + "..." if len(benchmark) > 120 else benchmark,
            "dok": s.get("dok"),
            "topics": s.get("topics", []),
        })

    return {
        "subject": config.get('subject', 'unknown'),
        "state": config.get('state', 'unknown'),
        "grade_level": config.get('grade_level', 'unknown'),
        "total_count": len(compact),
        "standards": compact,
    }


# ═══════════════════════════════════════════════════════
# RESOURCE TOOLS
# ═══════════════════════════════════════════════════════

def _extract_pdf_text(filepath):
    """Extract text from a PDF file path using PyMuPDF."""
    try:
        import fitz
        doc = fitz.open(filepath)
        pages = []
        for page in doc:
            pages.append(page.get_text())
        page_count = len(pages)
        doc.close()
        return "\n\n".join(pages), page_count
    except ImportError:
        return "[PDF extraction requires PyMuPDF: pip install pymupdf]", 0
    except Exception as e:
        return f"[Error extracting PDF: {e}]", 0


def _extract_docx_text(filepath):
    """Extract text from a DOCX file path using python-docx."""
    try:
        from docx import Document
        from docx.text.paragraph import Paragraph
        from docx.table import Table

        doc = Document(filepath)
        full_text = []
        for element in doc.element.body:
            if element.tag.endswith('p'):
                para = Paragraph(element, doc)
                if para.text.strip():
                    full_text.append(para.text)
            elif element.tag.endswith('tbl'):
                table = Table(element, doc)
                for row in table.rows:
                    row_text = [cell.text.strip() for cell in row.cells if cell.text.strip()]
                    if row_text:
                        full_text.append(' | '.join(row_text))
        return '\n'.join(full_text)
    except ImportError:
        return "[DOCX extraction requires python-docx: pip install python-docx]"
    except Exception as e:
        return f"[Error extracting DOCX: {e}]"


def list_resources_tool():
    """List all uploaded supporting documents from the documents directory."""
    if not os.path.isdir(DOCUMENTS_DIR):
        return {"documents": [], "message": "No documents directory found. Upload documents in Settings > Resources."}

    documents = []
    try:
        for fname in sorted(os.listdir(DOCUMENTS_DIR)):
            if fname.endswith('.meta.json'):
                continue
            fpath = os.path.join(DOCUMENTS_DIR, fname)
            if not os.path.isfile(fpath):
                continue

            # Try to load metadata
            meta_path = fpath + ".meta.json"
            meta = {}
            if os.path.exists(meta_path):
                try:
                    with open(meta_path, 'r', encoding='utf-8') as f:
                        meta = json.load(f)
                except Exception:
                    pass

            size_kb = round(os.path.getsize(fpath) / 1024, 1)
            documents.append({
                "filename": fname,
                "doc_type": meta.get("doc_type", "unknown"),
                "description": meta.get("description", ""),
                "size_kb": size_kb,
            })
    except Exception as e:
        return {"error": f"Error reading documents directory: {e}"}

    return {"documents": documents, "total": len(documents)}


MAX_RESOURCE_TEXT = 50000


def read_resource_tool(filename):
    """Read and return the text content of an uploaded document."""
    if not filename or not filename.strip():
        return {"error": "No filename provided"}

    # Prevent path traversal
    safe_name = os.path.basename(filename)
    filepath = os.path.join(DOCUMENTS_DIR, safe_name)

    if not os.path.exists(filepath):
        return {"error": f"Document not found: {safe_name}. Use list_resources to see available files."}

    ext = os.path.splitext(safe_name)[1].lower()
    pages = None

    try:
        if ext == '.pdf':
            content, pages = _extract_pdf_text(filepath)
        elif ext in ('.docx', '.doc'):
            content = _extract_docx_text(filepath)
        elif ext in ('.txt', '.md'):
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
        else:
            return {"error": f"Unsupported file type: {ext}. Supported: PDF, DOCX, DOC, TXT, MD."}
    except Exception as e:
        return {"error": f"Error reading {safe_name}: {e}"}

    truncated = False
    if len(content) > MAX_RESOURCE_TEXT:
        content = content[:MAX_RESOURCE_TEXT]
        truncated = True

    result = {
        "filename": safe_name,
        "content": content,
    }
    if pages:
        result["pages"] = pages
    if truncated:
        result["warning"] = f"Content truncated to {MAX_RESOURCE_TEXT} characters. Full document is larger."

    # Include metadata if available
    meta_path = filepath + ".meta.json"
    if os.path.exists(meta_path):
        try:
            with open(meta_path, 'r', encoding='utf-8') as f:
                meta = json.load(f)
            result["doc_type"] = meta.get("doc_type", "unknown")
            result["description"] = meta.get("description", "")
        except Exception:
            pass

    return result


# ═══════════════════════════════════════════════════════
# TOOL DISPATCH
# ═══════════════════════════════════════════════════════

TOOL_HANDLERS = {
    "query_grades": query_grades,
    "get_student_summary": get_student_summary,
    "get_class_analytics": get_class_analytics,
    "get_assignment_stats": get_assignment_stats,
    "list_assignments": list_assignments_tool,
    "create_focus_assignment": create_focus_assignment,
    "export_grades_csv": export_grades_csv,
    "analyze_grade_causes": analyze_grade_causes,
    "get_feedback_patterns": get_feedback_patterns,
    "compare_periods": compare_periods,
    "recommend_next_lesson": recommend_next_lesson,
    "lookup_student_info": lookup_student_info,
    "get_missing_assignments": get_missing_assignments,
    "generate_worksheet": generate_worksheet_tool,
    "generate_document": generate_document_tool,
    "generate_csv": generate_csv_tool,
    "save_document_style": save_document_style_tool,
    "list_document_styles": list_document_styles_tool,
    "save_memory": save_memory,
    "get_standards": get_standards_tool,
    "list_all_standards": list_all_standards_tool,
    "get_recent_lessons": get_recent_lessons,
    "get_calendar": get_calendar,
    "schedule_lesson": schedule_lesson_tool,
    "add_calendar_holiday": add_calendar_holiday,
    "list_resources": list_resources_tool,
    "read_resource": read_resource_tool,
}


def execute_tool(tool_name, tool_input):
    """Execute a tool by name with the given input."""
    handler = TOOL_HANDLERS.get(tool_name)
    if not handler:
        return {"error": f"Unknown tool: {tool_name}"}

    try:
        return handler(**tool_input)
    except Exception as e:
        return {"error": f"Tool execution error: {str(e)}"}
