"""
Settings-related API routes for Graider.
Handles rubric configuration, global settings, file uploads, and accommodations.
"""
import os
import json
import csv
import re
import subprocess
import threading
from datetime import datetime
from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename

# Import accommodation module
try:
    from backend.accommodations import (
        load_presets, save_preset, delete_preset,
        load_student_accommodations, set_student_accommodation,
        get_student_accommodation, remove_student_accommodation,
        import_accommodations_from_csv, clear_all_accommodations,
        get_accommodation_stats, export_student_accommodations,
        audit_log_accommodation
    )
except ImportError:
    from accommodations import (
        load_presets, save_preset, delete_preset,
        load_student_accommodations, set_student_accommodation,
        get_student_accommodation, remove_student_accommodation,
        import_accommodations_from_csv, clear_all_accommodations,
        get_accommodation_stats, export_student_accommodations,
        audit_log_accommodation
    )

settings_bp = Blueprint('settings', __name__)

# Data directories
GRAIDER_DATA_DIR = os.path.expanduser("~/.graider_data")
ROSTERS_DIR = os.path.join(GRAIDER_DATA_DIR, "rosters")
PERIODS_DIR = os.path.join(GRAIDER_DATA_DIR, "periods")
DOCUMENTS_DIR = os.path.join(GRAIDER_DATA_DIR, "documents")
PARENT_CONTACTS_FILE = os.path.join(GRAIDER_DATA_DIR, "parent_contacts.json")
EXPORTS_DIR = os.path.expanduser("~/.graider_exports")

# Ensure directories exist
for dir_path in [GRAIDER_DATA_DIR, ROSTERS_DIR, PERIODS_DIR, DOCUMENTS_DIR, EXPORTS_DIR]:
    os.makedirs(dir_path, exist_ok=True)

ALLOWED_CSV_EXTENSIONS = {'csv', 'xlsx', 'xls'}
ALLOWED_DOC_EXTENSIONS = {'pdf', 'docx', 'doc', 'txt', 'md'}


def parse_student_name(name):
    """Parse various student name formats and return (first, last)."""
    if not name:
        return '', ''

    name = name.strip()

    # Format: "Last; First Middle" (semicolon separator)
    if ';' in name:
        parts = name.split(';', 1)
        last = parts[0].strip()
        first_middle = parts[1].strip() if len(parts) > 1 else ''
        first = first_middle.split()[0] if first_middle else ''
        return first, last

    # Format: "Last, First Middle" (comma separator)
    if ',' in name:
        parts = name.split(',', 1)
        last = parts[0].strip()
        first_middle = parts[1].strip() if len(parts) > 1 else ''
        first = first_middle.split()[0] if first_middle else ''
        return first, last

    # Format: "First Last" or "First Middle Last" (space separated)
    parts = name.split()
    if len(parts) >= 2:
        first = parts[0]
        last = parts[-1]
        return first, last
    elif len(parts) == 1:
        return parts[0], ''

    return '', ''


def get_students_from_period_file(filepath):
    """Extract student list from a period CSV/Excel file."""
    students = []
    try:
        if filepath.endswith(('.xlsx', '.xls')):
            # Only import pandas for Excel files
            try:
                import pandas as pd
                df = pd.read_excel(filepath)
                name_cols = [c for c in df.columns if any(x in c.lower() for x in ['name', 'student', 'first', 'last'])]
                if name_cols:
                    first_col = next((c for c in name_cols if 'first' in c.lower()), None)
                    last_col = next((c for c in name_cols if 'last' in c.lower()), None)
                    if first_col and last_col:
                        for _, row in df.iterrows():
                            first = str(row[first_col]).strip() if pd.notna(row[first_col]) else ''
                            last = str(row[last_col]).strip() if pd.notna(row[last_col]) else ''
                            if first or last:
                                students.append({"first": first, "last": last, "full": f"{first} {last}".strip()})
                    else:
                        name_col = name_cols[0]
                        for _, row in df.iterrows():
                            name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
                            if name:
                                first, last = parse_student_name(name)
                                students.append({"first": first, "last": last, "full": name})
            except ImportError:
                print(f"Warning: pandas not installed, cannot read Excel file {filepath}")
        else:
            # Use built-in csv module for CSV files
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                first_col = next((h for h in reader.fieldnames if 'first' in h.lower()), None)
                last_col = next((h for h in reader.fieldnames if 'last' in h.lower()), None)
                name_col = next((h for h in reader.fieldnames if any(x in h.lower() for x in ['name', 'student'])), None)
                id_col = next((h for h in reader.fieldnames if h.lower().strip() == 'student id'), None)

                for row in reader:
                    student_id = row.get(id_col, '').strip() if id_col else ''
                    if first_col and last_col:
                        first = row.get(first_col, '').strip()
                        last = row.get(last_col, '').strip()
                        if first or last:
                            students.append({"first": first, "last": last, "full": f"{first} {last}".strip(), "id": student_id})
                    elif name_col:
                        name = row.get(name_col, '').strip()
                        if name:
                            first, last = parse_student_name(name)
                            students.append({"first": first, "last": last, "full": name, "id": student_id})
    except Exception as e:
        print(f"Error reading period file {filepath}: {e}")

    return students


@settings_bp.route('/api/save-rubric', methods=['POST'])
def save_rubric():
    """Save rubric configuration to JSON file."""
    data = request.json
    rubric_path = os.path.expanduser("~/.graider_rubric.json")

    try:
        with open(rubric_path, 'w') as f:
            json.dump(data, f, indent=2)
        return jsonify({"status": "saved"})
    except Exception as e:
        return jsonify({"error": str(e)})


@settings_bp.route('/api/load-rubric')
def load_rubric():
    """Load rubric configuration from JSON file."""
    rubric_path = os.path.expanduser("~/.graider_rubric.json")

    if not os.path.exists(rubric_path):
        return jsonify({"rubric": None})

    try:
        with open(rubric_path, 'r') as f:
            data = json.load(f)
        return jsonify({"rubric": data})
    except Exception as e:
        return jsonify({"error": str(e)})


@settings_bp.route('/api/save-global-settings', methods=['POST'])
def save_global_settings():
    """Save global AI notes and settings."""
    data = request.json
    settings_path = os.path.expanduser("~/.graider_settings.json")

    try:
        with open(settings_path, 'w') as f:
            json.dump(data, f, indent=2)
        return jsonify({"status": "saved"})
    except Exception as e:
        return jsonify({"error": str(e)})


@settings_bp.route('/api/load-global-settings')
def load_global_settings():
    """Load global AI notes and settings."""
    settings_path = os.path.expanduser("~/.graider_settings.json")

    if not os.path.exists(settings_path):
        return jsonify({"settings": None})

    try:
        with open(settings_path, 'r') as f:
            data = json.load(f)
        return jsonify({"settings": data})
    except Exception as e:
        return jsonify({"error": str(e)})


def allowed_file(filename, allowed_extensions):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


def parse_csv_headers(filepath):
    """Parse CSV and detect column mappings."""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            # Count rows
            row_count = sum(1 for _ in reader)
        return {'headers': headers, 'row_count': row_count}
    except Exception as e:
        return {'error': str(e)}


@settings_bp.route('/api/upload-roster', methods=['POST'])
def upload_roster():
    """Upload and process a roster CSV file."""
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    if not allowed_file(file.filename, ALLOWED_CSV_EXTENSIONS):
        return jsonify({"error": "Invalid file type. Use CSV, XLS, or XLSX"}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(ROSTERS_DIR, filename)
    file.save(filepath)

    # Parse and analyze the CSV
    result = parse_csv_headers(filepath)
    if 'error' in result:
        return jsonify({"error": result['error']}), 400

    # Save roster metadata
    metadata = {
        'filename': filename,
        'filepath': filepath,
        'headers': result['headers'],
        'row_count': result['row_count'],
        'column_mapping': {}  # Will be set by frontend
    }

    metadata_path = os.path.join(ROSTERS_DIR, f"{filename}.meta.json")
    with open(metadata_path, 'w') as f:
        json.dump(metadata, f, indent=2)

    return jsonify({
        "status": "uploaded",
        "filename": filename,
        "headers": result['headers'],
        "row_count": result['row_count']
    })


@settings_bp.route('/api/save-roster-mapping', methods=['POST'])
def save_roster_mapping():
    """Save column mapping for a roster file."""
    data = request.json
    filename = data.get('filename')
    mapping = data.get('mapping', {})

    if not filename:
        return jsonify({"error": "No filename provided"}), 400

    metadata_path = os.path.join(ROSTERS_DIR, f"{filename}.meta.json")
    if not os.path.exists(metadata_path):
        return jsonify({"error": "Roster metadata not found"}), 404

    with open(metadata_path, 'r') as f:
        metadata = json.load(f)

    metadata['column_mapping'] = mapping
    with open(metadata_path, 'w') as f:
        json.dump(metadata, f, indent=2)

    return jsonify({"status": "saved"})


@settings_bp.route('/api/list-rosters')
def list_rosters():
    """List all uploaded roster files."""
    rosters = []
    for f in os.listdir(ROSTERS_DIR):
        if f.endswith('.meta.json'):
            try:
                with open(os.path.join(ROSTERS_DIR, f), 'r') as mf:
                    metadata = json.load(mf)
                    rosters.append(metadata)
            except:
                pass
    return jsonify({"rosters": rosters})


@settings_bp.route('/api/delete-roster', methods=['POST'])
def delete_roster():
    """Delete a roster file."""
    data = request.json
    filename = data.get('filename')

    if not filename:
        return jsonify({"error": "No filename provided"}), 400

    filepath = os.path.join(ROSTERS_DIR, secure_filename(filename))
    metadata_path = os.path.join(ROSTERS_DIR, f"{secure_filename(filename)}.meta.json")

    try:
        if os.path.exists(filepath):
            os.remove(filepath)
        if os.path.exists(metadata_path):
            os.remove(metadata_path)
        return jsonify({"status": "deleted"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@settings_bp.route('/api/upload-period', methods=['POST'])
def upload_period():
    """Upload a period/class CSV file."""
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    period_name = request.form.get('period_name', 'Period')

    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    if not allowed_file(file.filename, ALLOWED_CSV_EXTENSIONS):
        return jsonify({"error": "Invalid file type. Use CSV, XLS, or XLSX"}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(PERIODS_DIR, filename)
    file.save(filepath)

    result = parse_csv_headers(filepath)
    if 'error' in result:
        return jsonify({"error": result['error']}), 400

    metadata = {
        'filename': filename,
        'filepath': filepath,
        'period_name': period_name,
        'headers': result['headers'],
        'row_count': result['row_count'],
        'column_mapping': {}
    }

    metadata_path = os.path.join(PERIODS_DIR, f"{filename}.meta.json")
    with open(metadata_path, 'w') as f:
        json.dump(metadata, f, indent=2)

    return jsonify({
        "status": "uploaded",
        "filename": filename,
        "period_name": period_name,
        "headers": result['headers'],
        "row_count": result['row_count']
    })


@settings_bp.route('/api/list-periods')
def list_periods():
    """List all uploaded period files with their students."""
    periods = []
    if os.path.exists(PERIODS_DIR):
        for f in os.listdir(PERIODS_DIR):
            if f.endswith('.meta.json'):
                try:
                    with open(os.path.join(PERIODS_DIR, f), 'r') as mf:
                        metadata = json.load(mf)
                        # Load students for this period
                        period_file = os.path.join(PERIODS_DIR, metadata.get('filename', ''))
                        if os.path.exists(period_file):
                            metadata['students'] = get_students_from_period_file(period_file)
                        else:
                            metadata['students'] = []
                        periods.append(metadata)
                except Exception as e:
                    print(f"Error loading period metadata {f}: {e}")
    return jsonify({"periods": periods})


@settings_bp.route('/api/delete-period', methods=['POST'])
def delete_period():
    """Delete a period file."""
    data = request.json
    filename = data.get('filename')

    if not filename:
        return jsonify({"error": "No filename provided"}), 400

    filepath = os.path.join(PERIODS_DIR, secure_filename(filename))
    metadata_path = os.path.join(PERIODS_DIR, f"{secure_filename(filename)}.meta.json")

    try:
        if os.path.exists(filepath):
            os.remove(filepath)
        if os.path.exists(metadata_path):
            os.remove(metadata_path)
        return jsonify({"status": "deleted"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@settings_bp.route('/api/update-period-level', methods=['POST'])
def update_period_level():
    """Update the class level (standard/advanced/support) for a period."""
    data = request.json
    filename = data.get('filename')
    class_level = data.get('class_level', 'standard')

    if not filename:
        return jsonify({"error": "No filename provided"}), 400

    if class_level not in ['standard', 'advanced', 'support']:
        return jsonify({"error": "Invalid class level. Use: standard, advanced, or support"}), 400

    metadata_path = os.path.join(PERIODS_DIR, f"{secure_filename(filename)}.meta.json")

    if not os.path.exists(metadata_path):
        return jsonify({"error": "Period metadata not found"}), 404

    try:
        with open(metadata_path, 'r') as f:
            metadata = json.load(f)

        metadata['class_level'] = class_level

        with open(metadata_path, 'w') as f:
            json.dump(metadata, f, indent=2)

        return jsonify({"status": "updated", "class_level": class_level})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@settings_bp.route('/api/get-period-students', methods=['POST'])
def get_period_students():
    """Get student names from a period CSV file."""
    data = request.json
    filename = data.get('filename')

    if not filename:
        return jsonify({"error": "No filename provided"}), 400

    filepath = os.path.join(PERIODS_DIR, secure_filename(filename))

    if not os.path.exists(filepath):
        return jsonify({"error": "Period file not found"}), 404

    try:
        students = get_students_from_period_file(filepath)
        return jsonify({"students": students, "count": len(students)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@settings_bp.route('/api/upload-document', methods=['POST'])
def upload_document():
    """Upload a supporting document for lesson planning/grading."""
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    doc_type = request.form.get('doc_type', 'general')  # 'curriculum', 'rubric', 'standards', 'general'
    description = request.form.get('description', '')

    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    if not allowed_file(file.filename, ALLOWED_DOC_EXTENSIONS):
        return jsonify({"error": "Invalid file type. Use PDF, DOCX, DOC, TXT, or MD"}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(DOCUMENTS_DIR, filename)
    file.save(filepath)

    metadata = {
        'filename': filename,
        'filepath': filepath,
        'doc_type': doc_type,
        'description': description,
        'size': os.path.getsize(filepath)
    }

    metadata_path = os.path.join(DOCUMENTS_DIR, f"{filename}.meta.json")
    with open(metadata_path, 'w') as f:
        json.dump(metadata, f, indent=2)

    return jsonify({
        "status": "uploaded",
        "filename": filename,
        "doc_type": doc_type
    })


@settings_bp.route('/api/list-documents')
def list_documents():
    """List all uploaded supporting documents."""
    documents = []
    for f in os.listdir(DOCUMENTS_DIR):
        if f.endswith('.meta.json'):
            try:
                with open(os.path.join(DOCUMENTS_DIR, f), 'r') as mf:
                    metadata = json.load(mf)
                    documents.append(metadata)
            except:
                pass
    return jsonify({"documents": documents})


@settings_bp.route('/api/delete-document', methods=['POST'])
def delete_document():
    """Delete a supporting document."""
    data = request.json
    filename = data.get('filename')

    if not filename:
        return jsonify({"error": "No filename provided"}), 400

    filepath = os.path.join(DOCUMENTS_DIR, secure_filename(filename))
    metadata_path = os.path.join(DOCUMENTS_DIR, f"{secure_filename(filename)}.meta.json")

    try:
        if os.path.exists(filepath):
            os.remove(filepath)
        if os.path.exists(metadata_path):
            os.remove(metadata_path)
        return jsonify({"status": "deleted"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
# PARENT CONTACTS IMPORT (from class_list Excel)
# ══════════════════════════════════════════════════════════════

def _is_email(value):
    """Detect if a string looks like an email address."""
    return bool(value and '@' in str(value))


def _clean_phone(value):
    """Return phone string or None."""
    if not value:
        return None
    s = str(value).strip()
    if '@' in s or not s:
        return None
    return s


def _find_header_row(ws, max_search=10, min_cols=3):
    """Find the first row in a worksheet that has at least min_cols non-empty values."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_search, values_only=True), start=1):
        if row:
            non_empty = sum(1 for v in row if v is not None and str(v).strip())
            if non_empty >= min_cols:
                return row_idx, [str(v).strip() if v is not None else '' for v in row]
    return None, []


def _suggest_mapping(headers, sample_rows):
    """Auto-suggest column mapping based on header names and sample data."""
    headers_lower = [h.lower() for h in headers]

    # Name column: first header containing "name" or "last" or "first"
    name_col = None
    for h, hl in zip(headers, headers_lower):
        if any(kw in hl for kw in ['last, first', 'last first', 'student name', 'name']):
            name_col = h
            break
    if not name_col:
        for h, hl in zip(headers, headers_lower):
            if 'last' in hl or 'first' in hl:
                name_col = h
                break

    # ID column: first header containing "id"
    id_col = None
    for h, hl in zip(headers, headers_lower):
        if 'student id' in hl or 'id' in hl:
            id_col = h
            break

    # Contact columns: headers containing contact-related keywords
    contact_cols = []
    for h, hl in zip(headers, headers_lower):
        if any(kw in hl for kw in ['contact', 'email', 'phone', 'cell', 'parent', 'guardian']):
            contact_cols.append(h)

    # Name format: check if header or sample data suggests "Last, First"
    name_format = 'first_last'
    if name_col:
        if 'last' in name_col.lower() and ('first' in name_col.lower() or ',' in name_col):
            name_format = 'last_first'
        elif sample_rows:
            col_idx = headers.index(name_col)
            for row in sample_rows[:5]:
                if col_idx < len(row) and row[col_idx]:
                    val = str(row[col_idx]).strip()
                    if ',' in val:
                        name_format = 'last_first'
                    break

    # ID strip digits: check if sample IDs are 9+ digits ending in grade-like suffix
    id_strip_digits = 0
    if id_col and sample_rows:
        col_idx = headers.index(id_col)
        grade_suffix_count = 0
        total_checked = 0
        for row in sample_rows[:10]:
            if col_idx < len(row) and row[col_idx]:
                try:
                    raw_id = str(int(float(str(row[col_idx]))))
                    if len(raw_id) >= 9:
                        total_checked += 1
                        suffix = raw_id[-2:]
                        if suffix.isdigit() and 6 <= int(suffix) <= 12:
                            grade_suffix_count += 1
                except (ValueError, TypeError):
                    pass
        if total_checked > 0 and grade_suffix_count / total_checked > 0.5:
            id_strip_digits = 2

    # Period column
    period_col = None
    for h, hl in zip(headers, headers_lower):
        if 'period' in hl or 'class' in hl or 'section' in hl:
            period_col = h
            break

    return {
        'name_col': name_col,
        'id_col': id_col,
        'contact_cols': contact_cols,
        'id_strip_digits': id_strip_digits,
        'name_format': name_format,
        'period_col': period_col,
    }


@settings_bp.route('/api/preview-parent-contacts', methods=['POST'])
def preview_parent_contacts():
    """
    Step 1: Upload class list file and return headers + suggested mapping.
    Supports xlsx (multi-sheet) and csv (single sheet).
    Does NOT process contacts — just previews the file structure.
    """
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    filename = file.filename.lower()
    if not filename.endswith(('.xlsx', '.xls', '.csv')):
        return jsonify({"error": "Please upload an Excel (.xlsx) or CSV file"}), 400

    try:
        ext = filename.rsplit('.', 1)[1]
        tmp_filename = "tmp_parent_contacts." + ext
        tmp_path = os.path.join(GRAIDER_DATA_DIR, tmp_filename)
        file.save(tmp_path)

        sheets = []
        all_headers = []
        all_sample_rows = []

        if ext in ('xlsx', 'xls'):
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header_row_idx, headers = _find_header_row(ws)
                if not headers:
                    continue
                # Filter out empty headers
                headers = [h for h in headers if h]

                # Get sample data rows for auto-suggest
                sample_rows = []
                if header_row_idx:
                    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 10, values_only=True):
                        sample_rows.append(list(row))

                # Count data rows
                row_count = 0
                if header_row_idx:
                    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                        if row and any(v is not None and str(v).strip() for v in row[:3]):
                            row_count += 1

                sheets.append({
                    'name': sheet_name,
                    'headers': headers,
                    'row_count': row_count,
                })
                if not all_headers:
                    all_headers = headers
                    all_sample_rows = sample_rows

            wb.close()

        else:
            # CSV
            with open(tmp_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                raw_headers = next(reader, [])
                headers = [h.strip() for h in raw_headers if h.strip()]
                sample_rows = []
                row_count = 0
                for row in reader:
                    if row and any(v.strip() for v in row[:3]):
                        row_count += 1
                        if len(sample_rows) < 10:
                            sample_rows.append(row)
            sheets.append({
                'name': 'Sheet1',
                'headers': headers,
                'row_count': row_count,
            })
            all_headers = headers
            all_sample_rows = sample_rows

        if not sheets:
            return jsonify({"error": "No data found in file"}), 400

        suggested = _suggest_mapping(all_headers, all_sample_rows)

        return jsonify({
            "filename": os.path.basename(tmp_path),
            "file_type": ext,
            "sheets": sheets,
            "suggested_mapping": suggested,
        })

    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@settings_bp.route('/api/save-parent-contact-mapping', methods=['POST'])
def save_parent_contact_mapping():
    """
    Step 2: Process the uploaded file using the teacher's confirmed column mapping.
    Reads the temp file, extracts contacts per the mapping, saves parent_contacts.json.
    """
    data = request.json
    if not data:
        return jsonify({"error": "No mapping data provided"}), 400

    name_col = data.get('name_col')
    name_format = data.get('name_format', 'last_first')
    id_col = data.get('id_col')
    id_strip_digits = int(data.get('id_strip_digits', 0))
    contact_cols = data.get('contact_cols', [])
    period_col = data.get('period_col')

    if not name_col:
        return jsonify({"error": "Name column is required"}), 400

    # Find the temp file
    tmp_path = None
    for ext in ('xlsx', 'xls', 'csv'):
        candidate = os.path.join(GRAIDER_DATA_DIR, "tmp_parent_contacts." + ext)
        if os.path.exists(candidate):
            tmp_path = candidate
            break

    if not tmp_path:
        return jsonify({"error": "No uploaded file found. Please upload again."}), 404

    try:
        contacts = {}
        total_students = 0
        period_counts = {}

        file_ext = tmp_path.rsplit('.', 1)[1].lower()

        def _process_rows(headers, rows_iter, default_period):
            """Process data rows using the confirmed mapping."""
            nonlocal total_students

            # Build column index map
            col_map = {}
            for i, h in enumerate(headers):
                col_map[h] = i

            name_idx = col_map.get(name_col)
            id_idx = col_map.get(id_col) if id_col else None
            contact_idxs = [col_map[c] for c in contact_cols if c in col_map]
            period_idx = col_map.get(period_col) if period_col else None

            if name_idx is None:
                return

            students_in_section = 0

            for row in rows_iter:
                if not row or name_idx >= len(row):
                    continue

                name_cell = row[name_idx]
                if not name_cell or not str(name_cell).strip():
                    continue

                name_str = str(name_cell).strip()

                # Parse name based on format
                if name_format == 'last_first':
                    if ',' in name_str:
                        parts = name_str.split(',', 1)
                        last_name = parts[0].strip()
                        first_middle = parts[1].strip() if len(parts) > 1 else ''
                    elif ';' in name_str:
                        parts = name_str.split(';', 1)
                        last_name = parts[0].strip()
                        first_middle = parts[1].strip() if len(parts) > 1 else ''
                    else:
                        last_name = name_str
                        first_middle = ''
                    student_name = (first_middle + ' ' + last_name).strip() if first_middle else last_name
                elif name_format == 'first_last':
                    parts = name_str.split()
                    if len(parts) >= 2:
                        first_middle = ' '.join(parts[:-1])
                        last_name = parts[-1]
                    else:
                        first_middle = name_str
                        last_name = ''
                    student_name = name_str
                else:
                    # single name
                    student_name = name_str
                    first_middle = name_str
                    last_name = ''

                # Student ID
                raw_id = ''
                if id_idx is not None and id_idx < len(row) and row[id_idx]:
                    try:
                        raw_id = str(int(float(str(row[id_idx]))))
                    except (ValueError, TypeError):
                        raw_id = str(row[id_idx]).strip()

                roster_id = raw_id
                if id_strip_digits > 0 and len(raw_id) > id_strip_digits:
                    roster_id = raw_id[:-id_strip_digits]

                # Use student name as key if no ID
                key = roster_id if roster_id else student_name

                # Scan contact columns
                emails = set()
                phones = []
                for ci in contact_idxs:
                    if ci < len(row) and row[ci]:
                        val_str = str(row[ci]).strip()
                        if _is_email(val_str):
                            emails.add(val_str.lower())
                        elif val_str:
                            phone = _clean_phone(val_str)
                            if phone and phone not in phones:
                                phones.append(phone)

                # Period
                period = default_period
                if period_idx is not None and period_idx < len(row) and row[period_idx]:
                    period = str(row[period_idx]).strip()

                total_students += 1
                students_in_section += 1

                if key in contacts:
                    existing = contacts[key]
                    existing['parent_emails'] = sorted(
                        set(existing['parent_emails']) | emails
                    )
                    for p in phones:
                        if p not in existing['parent_phones']:
                            existing['parent_phones'].append(p)
                else:
                    contacts[key] = {
                        'student_name': student_name,
                        'period': period,
                        'parent_emails': sorted(emails),
                        'parent_phones': phones,
                    }

            return students_in_section

        if file_ext in ('xlsx', 'xls'):
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header_row_idx, headers = _find_header_row(ws)
                if not headers or name_col not in headers:
                    continue

                rows = list(ws.iter_rows(min_row=header_row_idx + 1, values_only=True))
                count = _process_rows(headers, rows, sheet_name)
                if count:
                    period_counts[sheet_name] = count

            wb.close()
        else:
            # CSV
            with open(tmp_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                raw_headers = next(reader, [])
                headers = [h.strip() for h in raw_headers]
                rows = list(reader)
                count = _process_rows(headers, rows, 'Sheet1')
                if count:
                    period_counts['Sheet1'] = count

        # Clean up temp file
        try:
            os.remove(tmp_path)
        except OSError:
            pass

        # Save to JSON (same format as before)
        with open(PARENT_CONTACTS_FILE, 'w') as f:
            json.dump(contacts, f, indent=2)

        with_email = sum(1 for c in contacts.values() if c.get('parent_emails'))
        without_email = sum(1 for c in contacts.values() if not c.get('parent_emails'))

        return jsonify({
            "status": "imported",
            "total_students": total_students,
            "unique_students": len(contacts),
            "with_email": with_email,
            "without_email": without_email,
            "periods": period_counts,
        })

    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@settings_bp.route('/api/parent-contacts')
def get_parent_contacts():
    """Return stored parent contacts with summary stats."""
    if not os.path.exists(PARENT_CONTACTS_FILE):
        return jsonify({"contacts": {}, "count": 0, "with_email": 0})

    try:
        with open(PARENT_CONTACTS_FILE, 'r') as f:
            contacts = json.load(f)

        with_email = sum(1 for c in contacts.values() if c.get('parent_emails'))

        # Group by period
        period_stats = {}
        for c in contacts.values():
            period = c.get('period', 'Unknown')
            if period not in period_stats:
                period_stats[period] = {'total': 0, 'with_email': 0}
            period_stats[period]['total'] += 1
            if c.get('parent_emails'):
                period_stats[period]['with_email'] += 1

        return jsonify({
            "contacts": contacts,
            "count": len(contacts),
            "with_email": with_email,
            "without_email": len(contacts) - with_email,
            "period_stats": period_stats,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
# ACCOMMODATION MANAGEMENT (FERPA COMPLIANT)
# ══════════════════════════════════════════════════════════════
#
# FERPA Compliance Notes:
# - All accommodation data stored locally only (~/.graider_data/accommodations/)
# - Student IDs are never sent to AI - only accommodation TYPE
# - All access is audit logged
# - Data can be exported/deleted per FERPA requirements
# ══════════════════════════════════════════════════════════════

@settings_bp.route('/api/accommodation-presets')
def get_accommodation_presets():
    """Get all available accommodation presets (default + custom)."""
    presets = load_presets()
    return jsonify({"presets": list(presets.values())})


@settings_bp.route('/api/accommodation-presets', methods=['POST'])
def create_accommodation_preset():
    """Create or update a custom accommodation preset."""
    data = request.json
    if not data.get('name') or not data.get('ai_instructions'):
        return jsonify({"error": "Name and AI instructions required"}), 400

    if save_preset(data):
        return jsonify({"status": "saved", "preset": data})
    else:
        return jsonify({"error": "Failed to save preset"}), 500


@settings_bp.route('/api/accommodation-presets/<preset_id>', methods=['DELETE'])
def delete_accommodation_preset(preset_id):
    """Delete a custom accommodation preset."""
    if delete_preset(preset_id):
        return jsonify({"status": "deleted"})
    else:
        return jsonify({"error": "Cannot delete default presets or preset not found"}), 400


@settings_bp.route('/api/student-accommodations')
def get_all_student_accommodations():
    """
    Get all student accommodation mappings.
    FERPA: Returns student IDs with their accommodation settings.
    Data is stored and displayed locally only.
    """
    mappings = load_student_accommodations()
    presets = load_presets()

    # Build student ID → name lookup from period CSVs for name resolution
    id_to_name = {}
    if os.path.exists(PERIODS_DIR):
        for fname in os.listdir(PERIODS_DIR):
            if fname.endswith(('.csv', '.xlsx', '.xls')) and not fname.startswith('.'):
                try:
                    students = get_students_from_period_file(os.path.join(PERIODS_DIR, fname))
                    for s in students:
                        sid = s.get("id", "")
                        if sid:
                            name = s.get("full") or ((s.get("first", "") + " " + s.get("last", "")).strip())
                            if name:
                                id_to_name[sid] = name
                except Exception:
                    pass

    # Enrich with preset details for display
    enriched = {}
    for student_id, data in mappings.items():
        preset_details = []
        for preset_id in data.get("presets", []):
            if preset_id in presets:
                preset_details.append({
                    "id": preset_id,
                    "name": presets[preset_id].get("name", preset_id),
                    "icon": presets[preset_id].get("icon", "FileText")
                })

        # Resolve name: stored name > period CSV lookup > empty
        student_name = data.get("student_name", "") or id_to_name.get(student_id, "")

        enriched[student_id] = {
            "presets": preset_details,
            "custom_notes": data.get("custom_notes", ""),
            "student_name": student_name,
            "updated": data.get("updated", "")
        }

    return jsonify({"accommodations": enriched, "count": len(enriched)})


@settings_bp.route('/api/student-accommodations/<student_id>', methods=['GET'])
def get_single_student_accommodation(student_id):
    """Get accommodation settings for a specific student."""
    accommodation = get_student_accommodation(student_id)
    if accommodation:
        return jsonify({"accommodation": accommodation})
    else:
        return jsonify({"accommodation": None})


@settings_bp.route('/api/student-accommodations/<student_id>', methods=['POST'])
def set_single_student_accommodation(student_id):
    """Set accommodation presets for a student."""
    data = request.json
    preset_ids = data.get('presets', [])
    custom_notes = data.get('custom_notes', '')
    student_name = data.get('student_name', '')

    if set_student_accommodation(student_id, preset_ids, custom_notes, student_name):
        audit_log_accommodation("API_SET", f"Set accommodation for {student_id[:6]}...")
        return jsonify({"status": "saved"})
    else:
        return jsonify({"error": "Failed to save accommodation"}), 500


@settings_bp.route('/api/student-accommodations/<student_id>', methods=['DELETE'])
def delete_student_accommodation(student_id):
    """Remove accommodation settings for a student."""
    if remove_student_accommodation(student_id):
        return jsonify({"status": "deleted"})
    else:
        return jsonify({"error": "Student not found or deletion failed"}), 404


@settings_bp.route('/api/import-accommodations', methods=['POST'])
def import_accommodations():
    """
    Import accommodations from CSV file.
    Expected columns: student_id, accommodation_type, accommodation_notes (optional)
    """
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    id_col = request.form.get('id_column', 'student_id')
    accommodation_col = request.form.get('accommodation_column', 'accommodation_type')
    notes_col = request.form.get('notes_column', 'accommodation_notes')

    try:
        # Parse CSV
        content = file.read().decode('utf-8')
        reader = csv.DictReader(content.splitlines())
        csv_data = list(reader)

        result = import_accommodations_from_csv(csv_data, id_col, accommodation_col, notes_col)

        return jsonify({
            "status": "imported",
            "imported": result["imported"],
            "skipped": result["skipped"],
            "total": result["total"]
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@settings_bp.route('/api/export-accommodations')
def export_accommodations():
    """
    Export all accommodation data for backup.
    FERPA: Supports data portability requirements.
    """
    data = export_student_accommodations()
    return jsonify({
        "accommodations": data,
        "exported_at": __import__('datetime').datetime.now().isoformat()
    })


@settings_bp.route('/api/clear-accommodations', methods=['POST'])
def clear_accommodations():
    """
    Delete all student accommodation data.
    FERPA: Supports data deletion requirements.
    """
    if clear_all_accommodations():
        return jsonify({"status": "cleared"})
    else:
        return jsonify({"error": "Failed to clear accommodations"}), 500


@settings_bp.route('/api/accommodation-stats')
def accommodation_stats():
    """Get statistics about accommodation usage."""
    stats = get_accommodation_stats()
    return jsonify(stats)


# ============ API Keys Management ============

API_KEYS_FILE = os.path.join(GRAIDER_DATA_DIR, ".api_keys.json")


def load_api_keys():
    """Load API keys from secure storage."""
    if os.path.exists(API_KEYS_FILE):
        try:
            with open(API_KEYS_FILE, 'r') as f:
                return json.load(f)
        except:
            pass
    return {}


def save_api_keys_to_file(keys):
    """Save API keys to secure storage."""
    # Set restrictive permissions on the file
    with open(API_KEYS_FILE, 'w') as f:
        json.dump(keys, f)
    os.chmod(API_KEYS_FILE, 0o600)  # Owner read/write only


@settings_bp.route('/api/save-api-keys', methods=['POST'])
def save_api_keys():
    """Save API keys securely."""
    data = request.json
    openai_key = data.get('openai_key')
    anthropic_key = data.get('anthropic_key')
    gemini_key = data.get('gemini_key')

    # Load existing keys
    keys = load_api_keys()

    # Update keys if provided
    if openai_key:
        keys['openai'] = openai_key
    if anthropic_key:
        keys['anthropic'] = anthropic_key
    if gemini_key:
        keys['gemini'] = gemini_key

    # Save to file
    save_api_keys_to_file(keys)

    # Also update .env file for immediate use
    env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), '.env')
    env_lines = []
    if os.path.exists(env_path):
        with open(env_path, 'r') as f:
            env_lines = f.readlines()

    # Update or add keys in .env
    openai_found = False
    anthropic_found = False
    gemini_found = False
    new_lines = []
    for line in env_lines:
        if line.startswith('OPENAI_API_KEY=') and openai_key:
            new_lines.append(f'OPENAI_API_KEY={openai_key}\n')
            openai_found = True
        elif line.startswith('ANTHROPIC_API_KEY=') and anthropic_key:
            new_lines.append(f'ANTHROPIC_API_KEY={anthropic_key}\n')
            anthropic_found = True
        elif line.startswith('GEMINI_API_KEY=') and gemini_key:
            new_lines.append(f'GEMINI_API_KEY={gemini_key}\n')
            gemini_found = True
        else:
            new_lines.append(line)

    if openai_key and not openai_found:
        new_lines.append(f'OPENAI_API_KEY={openai_key}\n')
    if anthropic_key and not anthropic_found:
        new_lines.append(f'ANTHROPIC_API_KEY={anthropic_key}\n')
    if gemini_key and not gemini_found:
        new_lines.append(f'GEMINI_API_KEY={gemini_key}\n')

    with open(env_path, 'w') as f:
        f.writelines(new_lines)

    return jsonify({
        "status": "success",
        "openai_configured": bool(keys.get('openai')),
        "anthropic_configured": bool(keys.get('anthropic')),
        "gemini_configured": bool(keys.get('gemini'))
    })


@settings_bp.route('/api/check-api-keys')
def check_api_keys():
    """Check which API keys are configured (without exposing the keys)."""
    keys = load_api_keys()

    # Also check .env file
    env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), '.env')
    openai_in_env = False
    anthropic_in_env = False
    gemini_in_env = False

    if os.path.exists(env_path):
        with open(env_path, 'r') as f:
            content = f.read()
            openai_in_env = 'OPENAI_API_KEY=' in content and 'your-key-here' not in content
            anthropic_in_env = 'ANTHROPIC_API_KEY=' in content
            gemini_in_env = 'GEMINI_API_KEY=' in content

    return jsonify({
        "openai_configured": bool(keys.get('openai')) or openai_in_env,
        "anthropic_configured": bool(keys.get('anthropic')) or anthropic_in_env,
        "gemini_configured": bool(keys.get('gemini')) or gemini_in_env
    })


# ══════════════════════════════════════════════════════════════
# FOCUS ROSTER IMPORT
# ══════════════════════════════════════════════════════════════

FOCUS_IMPORT_FILE = os.path.join(GRAIDER_DATA_DIR, "focus_roster_import.json")

# Module-level state for Focus import process
_focus_import_state = {
    "status": "idle",
    "progress": "",
    "result": None,
    "error": None,
}


def _run_focus_import():
    """Run the Focus roster import script in a background thread."""
    global _focus_import_state
    _focus_import_state["status"] = "running"
    _focus_import_state["progress"] = "Starting Focus import..."
    _focus_import_state["result"] = None
    _focus_import_state["error"] = None

    script_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        "focus-roster-import.js"
    )

    if not os.path.exists(script_path):
        _focus_import_state["status"] = "failed"
        _focus_import_state["error"] = "focus-roster-import.js not found"
        return

    try:
        proc = subprocess.Popen(
            ["node", script_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )

        # Read stdout lines for progress updates
        for line in proc.stdout:
            line = line.strip()
            if not line:
                continue
            try:
                msg = json.loads(line)
                msg_type = msg.get("type", "")
                message = msg.get("message", "")

                if msg_type == "progress":
                    _focus_import_state["progress"] = message
                elif msg_type == "status":
                    _focus_import_state["progress"] = message
                elif msg_type == "complete":
                    _focus_import_state["progress"] = message
                elif msg_type == "error":
                    _focus_import_state["error"] = message
                elif msg_type == "warning":
                    _focus_import_state["progress"] = message
            except json.JSONDecodeError:
                pass

        proc.wait()

        if proc.returncode != 0:
            stderr = proc.stderr.read()
            if not _focus_import_state["error"]:
                _focus_import_state["error"] = stderr or "Import process exited with error"
            _focus_import_state["status"] = "failed"
            return

        # Process the import output file
        if not os.path.exists(FOCUS_IMPORT_FILE):
            _focus_import_state["status"] = "failed"
            _focus_import_state["error"] = "Import file not created"
            return

        with open(FOCUS_IMPORT_FILE, 'r') as f:
            import_data = json.load(f)

        result = _process_focus_import(import_data)
        _focus_import_state["status"] = "completed"
        _focus_import_state["result"] = result

    except Exception as e:
        _focus_import_state["status"] = "failed"
        _focus_import_state["error"] = str(e)


def _process_focus_import(import_data):
    """Process Focus import data: write period CSVs and update parent contacts.

    Handles enriched format from focus-roster-import.js:
    - periods keyed by "Period N" with course_codes, period_num, students
    - students have contacts (primary/secondary/third), schedule, has_504
    - backward-compatible parent_emails/parent_phones arrays
    """
    periods_data = import_data.get("periods", {})
    total_students = 0
    total_contacts = 0
    period_summary = {}

    # Load existing parent contacts (preserve non-Focus entries)
    contacts = {}
    if os.path.exists(PARENT_CONTACTS_FILE):
        try:
            with open(PARENT_CONTACTS_FILE, 'r') as f:
                contacts = json.load(f)
        except (json.JSONDecodeError, IOError):
            pass

    # Store student schedules separately for cross-teacher lookup
    schedules_path = os.path.join(GRAIDER_DATA_DIR, "student_schedules.json")

    for period_name, period_data in periods_data.items():
        students = period_data.get("students", [])
        if not students:
            continue

        course_codes = period_data.get("course_codes", [])
        period_num = period_data.get("period_num", 0)

        # Write period CSV
        safe_name = re.sub(r'[^\w\s-]', '', period_name).strip().replace(' ', '_')
        csv_filename = safe_name + ".csv"
        csv_path = os.path.join(PERIODS_DIR, csv_filename)

        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Student", "Student ID", "Local ID", "Grade", "Local Student ID", "Team"])
            for s in students:
                writer.writerow([
                    s.get("name", ""),
                    s.get("student_id", ""),
                    s.get("local_id", ""),
                    s.get("grade", ""),
                    s.get("local_id", ""),
                    ""
                ])

        # Write .meta.json with course code info
        meta = {
            "filename": csv_filename,
            "filepath": csv_path,
            "period_name": period_name,
            "period_num": period_num,
            "course_codes": course_codes,
            "headers": ["Student", "Student ID", "Local ID", "Grade", "Local Student ID", "Team"],
            "row_count": len(students),
            "column_mapping": {},
            "imported_from": "focus",
            "imported_at": import_data.get("imported_at", "")
        }
        meta_path = os.path.join(PERIODS_DIR, csv_filename + ".meta.json")
        with open(meta_path, 'w') as f:
            json.dump(meta, f, indent=2)

        # Update parent contacts with enriched data
        period_contacts = 0
        for s in students:
            sid = s.get("student_id", "")
            if not sid:
                continue

            emails = s.get("parent_emails", [])
            phones = s.get("parent_phones", [])
            rich_contacts = s.get("contacts", {})
            schedule = s.get("schedule", [])
            has_504 = s.get("has_504", False)

            contact_entry = {
                "student_name": s.get("name", ""),
                "period": period_name,
                "has_504": has_504,
                "parent_emails": sorted(set(emails)) if emails else [],
                "parent_phones": phones or [],
                "contacts": [],
                "schedule": schedule,
            }

            # Build structured contacts list from primary/secondary/third
            for role in ["primary", "secondary", "third"]:
                c = rich_contacts.get(role)
                if c:
                    contact_entry["contacts"].append({
                        "role": role,
                        "first_name": c.get("first_name", ""),
                        "last_name": c.get("last_name", ""),
                        "relationship": c.get("relationship", ""),
                        "phone": c.get("phone", ""),
                        "email": c.get("email", ""),
                        "call_out": c.get("call_out", False),
                    })

            # Merge with existing entry if present (preserve manually added data)
            if sid in contacts:
                existing = contacts[sid]
                # Merge emails
                merged_emails = sorted(
                    set(existing.get("parent_emails", [])) | set(contact_entry["parent_emails"])
                )
                contact_entry["parent_emails"] = merged_emails
                # Merge phones
                for p in existing.get("parent_phones", []):
                    if p not in contact_entry["parent_phones"]:
                        contact_entry["parent_phones"].append(p)

            contacts[sid] = contact_entry
            if emails or phones:
                period_contacts += 1

        total_students += len(students)
        total_contacts += period_contacts
        period_summary[period_name] = {
            "students": len(students),
            "course_codes": course_codes,
        }

    # Save updated parent contacts
    with open(PARENT_CONTACTS_FILE, 'w') as f:
        json.dump(contacts, f, indent=2)

    return {
        "periods_imported": len(period_summary),
        "total_students": total_students,
        "total_contacts": total_contacts,
        "period_summary": period_summary,
    }


@settings_bp.route('/api/import-from-focus', methods=['POST'])
def import_from_focus():
    """Trigger Focus SIS roster import via Playwright."""
    # Check credentials exist
    creds_path = os.path.join(GRAIDER_DATA_DIR, "portal_credentials.json")
    if not os.path.exists(creds_path):
        return jsonify({"error": "VPortal credentials not configured. Go to Settings > Tools > District Portal."}), 400

    if _focus_import_state["status"] == "running":
        return jsonify({"error": "Import already in progress"}), 409

    # Start import in background thread
    thread = threading.Thread(target=_run_focus_import, daemon=True)
    thread.start()

    return jsonify({"status": "started", "message": "Focus import started. A browser window will open for 2FA."})


@settings_bp.route('/api/focus-import-status')
def focus_import_status():
    """Get current status of the Focus import process."""
    return jsonify({
        "status": _focus_import_state["status"],
        "progress": _focus_import_state["progress"],
        "result": _focus_import_state["result"],
        "error": _focus_import_state["error"],
    })


# ══════════════════════════════════════════════════════════════
# INLINE ROSTER EDITOR (Add/Remove/Update Students)
# ══════════════════════════════════════════════════════════════

def _read_period_csv(filepath):
    """Read a period CSV and return (headers, rows) where rows are lists."""
    headers = []
    rows = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        for row in reader:
            rows.append(row)
    return headers, rows


def _write_period_csv(filepath, headers, rows):
    """Write headers + rows back to a period CSV."""
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)


def _update_meta_row_count(filename, new_count):
    """Update the row_count in a period's .meta.json file."""
    meta_path = os.path.join(PERIODS_DIR, secure_filename(filename) + ".meta.json")
    if os.path.exists(meta_path):
        try:
            with open(meta_path, 'r') as f:
                meta = json.load(f)
            meta['row_count'] = new_count
            with open(meta_path, 'w') as f:
                json.dump(meta, f, indent=2)
        except (json.JSONDecodeError, IOError):
            pass


def _find_student_id_col(headers):
    """Find the index of the Student ID column."""
    for i, h in enumerate(headers):
        if h.strip().lower() == 'student id':
            return i
    return -1


def _find_student_name_col(headers):
    """Find the index of the Student/Name column."""
    for i, h in enumerate(headers):
        hl = h.strip().lower()
        if hl in ('student', 'name', 'student name'):
            return i
    return 0  # default to first column


def _load_parent_contacts():
    """Load parent_contacts.json safely."""
    if os.path.exists(PARENT_CONTACTS_FILE):
        try:
            with open(PARENT_CONTACTS_FILE, 'r') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {}


def _save_parent_contacts(contacts):
    """Save parent_contacts.json."""
    with open(PARENT_CONTACTS_FILE, 'w') as f:
        json.dump(contacts, f, indent=2)


@settings_bp.route('/api/add-student', methods=['POST'])
def add_student():
    """Add a student to a period CSV and optionally to parent contacts."""
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    filename = data.get('period_filename')
    student_name = data.get('student_name', '').strip()
    student_id = data.get('student_id', '').strip()

    if not filename:
        return jsonify({"error": "period_filename is required"}), 400
    if not student_name:
        return jsonify({"error": "student_name is required"}), 400

    filepath = os.path.join(PERIODS_DIR, secure_filename(filename))
    if not os.path.exists(filepath):
        return jsonify({"error": "Period file not found"}), 404

    try:
        headers, rows = _read_period_csv(filepath)

        # Check for duplicate student ID
        if student_id:
            id_col = _find_student_id_col(headers)
            if id_col >= 0:
                for row in rows:
                    if id_col < len(row) and row[id_col].strip() == student_id:
                        return jsonify({"error": "Student ID already exists in this period"}), 409

        # Build new row matching existing headers
        local_id = data.get('local_id', '')
        grade = data.get('grade', '')
        new_row = []
        for h in headers:
            hl = h.strip().lower()
            if hl in ('student', 'name', 'student name'):
                new_row.append(student_name)
            elif hl == 'student id':
                new_row.append(student_id)
            elif hl == 'local id' or hl == 'local student id':
                new_row.append(local_id)
            elif hl == 'grade':
                new_row.append(grade)
            else:
                new_row.append('')

        rows.append(new_row)
        _write_period_csv(filepath, headers, rows)
        _update_meta_row_count(filename, len(rows))

        # Update parent contacts if provided
        parent_emails = data.get('parent_emails', [])
        parent_phones = data.get('parent_phones', [])
        if (parent_emails or parent_phones) and student_id:
            contacts = _load_parent_contacts()

            # Try to get the period_name from meta
            period_name = ''
            meta_path = os.path.join(PERIODS_DIR, secure_filename(filename) + ".meta.json")
            if os.path.exists(meta_path):
                try:
                    with open(meta_path, 'r') as f:
                        meta = json.load(f)
                    period_name = meta.get('period_name', '')
                except (json.JSONDecodeError, IOError):
                    pass

            contacts[student_id] = {
                "student_name": student_name,
                "period": period_name,
                "parent_emails": parent_emails if isinstance(parent_emails, list) else [e.strip() for e in parent_emails.split(',') if e.strip()],
                "parent_phones": parent_phones if isinstance(parent_phones, list) else [p.strip() for p in parent_phones.split(',') if p.strip()],
            }
            _save_parent_contacts(contacts)

        # Return updated student list
        students = get_students_from_period_file(filepath)
        return jsonify({"status": "added", "students": students, "count": len(students)})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@settings_bp.route('/api/remove-student', methods=['POST'])
def remove_student():
    """Remove a student from period CSV and parent contacts."""
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    filename = data.get('period_filename')
    student_id = data.get('student_id', '').strip()

    if not filename:
        return jsonify({"error": "period_filename is required"}), 400
    if not student_id:
        return jsonify({"error": "student_id is required"}), 400

    filepath = os.path.join(PERIODS_DIR, secure_filename(filename))
    if not os.path.exists(filepath):
        return jsonify({"error": "Period file not found"}), 404

    try:
        headers, rows = _read_period_csv(filepath)
        id_col = _find_student_id_col(headers)

        if id_col < 0:
            return jsonify({"error": "Could not find Student ID column in CSV"}), 400

        original_count = len(rows)
        rows = [row for row in rows if not (id_col < len(row) and row[id_col].strip() == student_id)]

        if len(rows) == original_count:
            return jsonify({"error": "Student not found in this period"}), 404

        _write_period_csv(filepath, headers, rows)
        _update_meta_row_count(filename, len(rows))

        # Remove from parent contacts
        contacts = _load_parent_contacts()
        if student_id in contacts:
            del contacts[student_id]
            _save_parent_contacts(contacts)

        students = get_students_from_period_file(filepath)
        return jsonify({"status": "removed", "students": students, "count": len(students)})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@settings_bp.route('/api/update-student', methods=['POST'])
def update_student():
    """Update a student's info in the period CSV and/or parent contacts."""
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    filename = data.get('period_filename')
    student_id = data.get('student_id', '').strip()

    if not filename:
        return jsonify({"error": "period_filename is required"}), 400
    if not student_id:
        return jsonify({"error": "student_id is required"}), 400

    filepath = os.path.join(PERIODS_DIR, secure_filename(filename))
    if not os.path.exists(filepath):
        return jsonify({"error": "Period file not found"}), 404

    try:
        new_name = data.get('student_name', '').strip()
        new_grade = data.get('grade', '').strip()
        parent_emails = data.get('parent_emails', [])
        parent_phones = data.get('parent_phones', [])

        # Update CSV if name or grade changed
        if new_name or new_grade:
            headers, rows = _read_period_csv(filepath)
            id_col = _find_student_id_col(headers)
            name_col = _find_student_name_col(headers)
            grade_col = -1
            for i, h in enumerate(headers):
                if h.strip().lower() == 'grade':
                    grade_col = i
                    break

            found = False
            for row in rows:
                if id_col >= 0 and id_col < len(row) and row[id_col].strip() == student_id:
                    if new_name and name_col >= 0 and name_col < len(row):
                        row[name_col] = new_name
                    if new_grade and grade_col >= 0 and grade_col < len(row):
                        row[grade_col] = new_grade
                    found = True
                    break

            if not found:
                return jsonify({"error": "Student not found in this period"}), 404

            _write_period_csv(filepath, headers, rows)

        # Update parent contacts
        contacts = _load_parent_contacts()
        if student_id in contacts:
            if parent_emails is not None:
                if isinstance(parent_emails, str):
                    parent_emails = [e.strip() for e in parent_emails.split(',') if e.strip()]
                contacts[student_id]["parent_emails"] = parent_emails
            if parent_phones is not None:
                if isinstance(parent_phones, str):
                    parent_phones = [p.strip() for p in parent_phones.split(',') if p.strip()]
                contacts[student_id]["parent_phones"] = parent_phones
            if new_name:
                contacts[student_id]["student_name"] = new_name
        else:
            # Create new contact entry
            period_name = ''
            meta_path = os.path.join(PERIODS_DIR, secure_filename(filename) + ".meta.json")
            if os.path.exists(meta_path):
                try:
                    with open(meta_path, 'r') as f:
                        meta = json.load(f)
                    period_name = meta.get('period_name', '')
                except (json.JSONDecodeError, IOError):
                    pass

            if isinstance(parent_emails, str):
                parent_emails = [e.strip() for e in parent_emails.split(',') if e.strip()]
            if isinstance(parent_phones, str):
                parent_phones = [p.strip() for p in parent_phones.split(',') if p.strip()]

            contacts[student_id] = {
                "student_name": new_name or "",
                "period": period_name,
                "parent_emails": parent_emails or [],
                "parent_phones": parent_phones or [],
            }

        _save_parent_contacts(contacts)

        return jsonify({"status": "updated"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500
